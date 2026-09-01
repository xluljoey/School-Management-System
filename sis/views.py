from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.models import User
from django.core.exceptions import PermissionDenied
from django.db import transaction
from django.db.models import Q, Count, Prefetch
from django.http import HttpResponse, JsonResponse
from django.urls import reverse
from django.views.decorators.http import require_POST
import calendar
import json
from datetime import date
from .models import (
    Parent, Student, Subject, SubjectAssessment, ClassRoom, Enrollment,
    StaffProfile, AcademicSession, Term, ClassSubject, PromotionCriteria,
    StaffClassSubject, Department, Designation, GradeVerification,
    MidTermRecord, Notification, Timetable, TimetableSlot,
)
from .forms import (
    ParentForm, StudentRegistrationForm, StaffRegistrationForm, EnrollmentForm,
    MarkSubmissionForm,
)

def _is_staff_or_admin(user):
    return user.is_active and (user.is_superuser or user.is_staff or hasattr(user, 'staff_profile'))


def _is_admin(user):
    return user.is_active and user.is_superuser


def get_ordinal(n):
    """Convert a number to ordinal format: 1 -> 1st, 2 -> 2nd, etc."""
    if n is None:
        return "—"
    n = int(n)
    if 10 <= n % 100 <= 20:
        suffix = 'th'
    else:
        suffix = {1: 'st', 2: 'nd', 3: 'rd'}.get(n % 10, 'th')
    return f"{n}{suffix}"


def is_jhs_class(classroom):
    """Check if a classroom is a JHS class."""
    return classroom and "jhs" in classroom.class_name.lower()


def get_grade_remarks_for_class(classroom):
    """Get appropriate GRADE_REMARKS based on class type."""
    if is_jhs_class(classroom):
        # WAEC standard for JHS
        return [
            (80, "A1", "Excellent"),
            (75, "A2", "Excellent"),
            (70, "B3", "Very Good"),
            (65, "C4", "Good"),
            (60, "C5", "Good"),
            (55, "C6", "Average"),
            (50, "D7", "Average"),
            (45, "E8", "Pass"),
            (0,  "F9", "Fail"),
        ]
    else:
        # Primary classes - keep letters but update remarks
        return [
            (80, "1", "Excellent"),
            (75, "2", "Excellent"),
            (70, "3", "Very Good"),
            (65, "4", "Good"),
            (60, "5", "Good"),
            (55, "6", "Average"),
            (50, "7", "Average"),
            (40, "8", "Pass"),
            (0,  "9", "Fail"),
        ]


def _build_dashboard_calendar_data(request, current_date=None):
    today = date.today()
    current_date = current_date or today
    year = current_date.year
    month = current_date.month
    month_name = current_date.strftime('%B %Y')

    active_timetables = Timetable.objects.filter(is_active=True)
    staff_profile = getattr(request.user, 'staff_profile', None)

    if request.user.is_superuser:
        pass
    elif staff_profile:
        assigned_class_ids = StaffClassSubject.objects.filter(staff=staff_profile).values_list('classroom_id', flat=True).distinct()
        active_timetables = active_timetables.filter(student_class_id__in=assigned_class_ids)
    else:
        active_timetables = active_timetables.none()

    active_timetables = active_timetables.select_related('student_class', 'academic_term').prefetch_related('slots__subject', 'slots__teacher')

    day_index = {'MON': 0, 'TUE': 1, 'WED': 2, 'THU': 3, 'FRI': 4, 'SAT': 5, 'SUN': 6}
    events_by_date = {}

    for timetable in active_timetables:
        for slot in timetable.slots.all():
            weekday_index = day_index.get(slot.day_of_week)
            if weekday_index is None:
                continue

            for week in calendar.Calendar(firstweekday=0).monthdatescalendar(year, month):
                for day in week:
                    if day.month != month:
                        continue
                    if day.weekday() != weekday_index:
                        continue

                    teacher_name = ''
                    if slot.teacher:
                        teacher_name = slot.teacher.user.get_full_name() or slot.teacher.user.username

                    event_entry = {
                        'date': day,
                        'subject': slot.subject.subject_name,
                        'start_time': slot.start_time.strftime('%H:%M'),
                        'teacher': teacher_name,
                        'class_name': timetable.student_class.class_name,
                    }
                    events_by_date.setdefault(day, []).append(event_entry)

    calendar_weeks = []
    for week in calendar.Calendar(firstweekday=0).monthdatescalendar(year, month):
        week_days = []
        for day in week:
            week_days.append({
                'date': day,
                'day': day.day if day.month == month else '',
                'is_current_month': day.month == month,
                'is_today': day == today and day.month == month and day.year == year,
                'events': events_by_date.get(day, [])[:2],
            })
        calendar_weeks.append(week_days)

    upcoming_events = []
    for day in sorted(events_by_date.keys()):
        for event in events_by_date[day]:
            upcoming_events.append(event)
    upcoming_events = sorted(upcoming_events, key=lambda item: (item['date'], item['start_time']))

    return {
        'month_label': month_name,
        'calendar_weeks': calendar_weeks,
        'upcoming_events': upcoming_events[:4],
        'has_calendar_events': bool(upcoming_events),
    }


# Create your views here.
@login_required
def dashboard_view(request):
    is_admin = request.user.is_superuser or request.user.groups.filter(name='Admin').exists()

    if is_admin:
        context = {
            'total_staff_count': StaffProfile.objects.count(),
            'environment': 'Academic Year Master Control',
        }
        return render(request, 'sis/admin_dashboard.html', context)

    if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.GET.get('calendar_partial') == '1':
        month_param = request.GET.get('month')
        year_param = request.GET.get('year')
        current_date = date.today()

        if month_param and year_param:
            try:
                current_date = date(int(year_param), int(month_param), 1)
            except ValueError:
                current_date = date.today()

        prev_month = current_date.month - 1 if current_date.month > 1 else 12
        prev_year = current_date.year if current_date.month > 1 else current_date.year - 1
        next_month = current_date.month + 1 if current_date.month < 12 else 1
        next_year = current_date.year if current_date.month < 12 else current_date.year + 1

        calendar_data = _build_dashboard_calendar_data(request, current_date)
        context = {
            'request': request,
            'calendar_month_label': calendar_data['month_label'],
            'calendar_weeks': calendar_data['calendar_weeks'],
            'upcoming_events': calendar_data['upcoming_events'],
            'has_calendar_events': calendar_data['has_calendar_events'],
            'calendar_year': current_date.year,
            'calendar_month': current_date.month,
            'calendar_prev_month': prev_month,
            'calendar_prev_year': prev_year,
            'calendar_next_month': next_month,
            'calendar_next_year': next_year,
        }
        return render(request, 'sis/partials/dashboard_calendar.html', context)

    total_students = Student.objects.count()
    total_staff = StaffProfile.objects.count()
    active_classes = ClassRoom.objects.count()
    current_session = AcademicSession.objects.filter(is_current=True).first()
    current_term = Term.objects.filter(is_active=True).first() if current_session else None

    total_boys = Student.objects.filter(gender__iexact='Male').count()
    total_girls = Student.objects.filter(gender__iexact='Female').count()
    boy_pct = int((total_boys / total_students) * 100) if total_students > 0 else 0
    girl_pct = int((total_girls / total_students) * 100) if total_students > 0 else 0

    staff_profile = getattr(request.user, 'staff_profile', None)

    if request.user.is_superuser:
        assigned_classes = ClassRoom.objects.all()
    elif staff_profile:
        assigned_class_ids = StaffClassSubject.objects.filter(staff=staff_profile).values_list('classroom_id', flat=True).distinct()
        assigned_classes = ClassRoom.objects.filter(id__in=assigned_class_ids) if assigned_class_ids else ClassRoom.objects.none()
    else:
        assigned_classes = ClassRoom.objects.none()

    staff_student_count = 0
    staff_subject_count = 0
    active_environment_string = ''

    if staff_profile and not request.user.is_superuser:
        if assigned_class_ids:
            staff_student_count = Student.objects.filter(enrollments__classroom_id__in=assigned_class_ids).distinct().count()
        staff_subject_count = Subject.objects.filter(assigned_teachers__staff=staff_profile).distinct().count()

    if current_term:
        term_name = current_term.term_name
    elif current_session:
        term_name = 'Term 1'
    else:
        term_name = ''
    active_environment_string = f"{term_name}, {current_session.academic_year}" if current_session and term_name else (str(current_session) if current_session else '')

    month_param = request.GET.get('month')
    year_param = request.GET.get('year')
    current_date = date.today()

    if month_param and year_param:
        try:
            current_date = date(int(year_param), int(month_param), 1)
        except ValueError:
            current_date = date.today()

    prev_month = current_date.month - 1 if current_date.month > 1 else 12
    prev_year = current_date.year if current_date.month > 1 else current_date.year - 1
    next_month = current_date.month + 1 if current_date.month < 12 else 1
    next_year = current_date.year if current_date.month < 12 else current_date.year + 1

    calendar_data = _build_dashboard_calendar_data(request, current_date)

    context = {
        'total_students': total_students,
        'total_staff': total_staff,
        'active_classes': active_classes,
        'current_session': current_session,
        'current_term': current_term,
        'total_boys': total_boys,
        'total_girls': total_girls,
        'boy_pct': boy_pct,
        'girl_pct': girl_pct,
        'staff_student_count': staff_student_count,
        'staff_subject_count': staff_subject_count,
        'active_environment_string': active_environment_string,
        'assigned_classes': assigned_classes,
        'calendar_month_label': calendar_data['month_label'],
        'calendar_weeks': calendar_data['calendar_weeks'],
        'upcoming_events': calendar_data['upcoming_events'],
        'has_calendar_events': calendar_data['has_calendar_events'],
        'calendar_year': current_date.year,
        'calendar_month': current_date.month,
        'calendar_prev_month': prev_month,
        'calendar_prev_year': prev_year,
        'calendar_next_month': next_month,
        'calendar_next_year': next_year,
    }
    return render(request, 'sis/staff_dashboard.html', context)


@login_required
def student_list_view(request):
    is_admin = _is_admin(request.user)

    if is_admin:
        students = Student.objects.filter(is_alumni=False).select_related('classroom')
    else:
        staff = request.user.staff_profile
        visible_class_ids = set(StaffClassSubject.objects.filter(
            staff=staff
        ).values_list('classroom_id', flat=True))
        if staff.form_class:
            visible_class_ids.add(staff.form_class_id)
        visible_class_ids.update(
            ClassRoom.objects.filter(form_master=staff).values_list('id', flat=True)
        )
        students = Student.objects.filter(
            classroom_id__in=visible_class_ids, is_alumni=False
        ).select_related('classroom').distinct()

    gender = request.GET.get('gender', '').strip()
    class_id = request.GET.get('class_id', '').strip()
    status = request.GET.get('status', '').strip()

    if gender and gender != 'all':
        students = students.filter(gender__iexact=gender)
    if class_id and class_id != 'all':
        students = students.filter(classroom_id=class_id)
    if status and status != 'all':
        students = students.filter(status__iexact=status)

    classrooms = ClassRoom.objects.all()
    return render(request, 'sis/student_list.html', {
        'students': students,
        'classrooms': classrooms,
        'is_admin': is_admin,
        'selected_gender': gender or 'all',
        'selected_class': class_id or 'all',
        'selected_status': status or 'all',
        'total_count': students.count(),
    })


@login_required
def alumni_list_view(request):
    if not _is_staff_or_admin(request.user):
        raise PermissionDenied

    students = Student.objects.filter(is_alumni=True).select_related('classroom').order_by('last_name', 'first_name')

    return render(request, 'sis/alumni_list.html', {
        'students': students,
        'total_count': students.count(),
    })


@login_required
def student_detail_view(request, student_id):
    student = get_object_or_404(Student, pk=student_id)
    return render(request, 'sis/student_detail.html', {'student': student})


@login_required
@user_passes_test(_is_admin)
def student_edit_view(request, student_id):
    student = get_object_or_404(Student, pk=student_id)
    if request.method == 'POST':
        student_form = StudentRegistrationForm(request.POST, request.FILES, instance=student)
        if student_form.is_valid():
            student_form.save()
            messages.success(request, 'Student updated successfully.')
            return redirect('student_detail', student_id=student.id)
    else:
        student_form = StudentRegistrationForm(instance=student)
    return render(request, 'sis/student_registration.html', {
        'student_form': student_form,
        'is_edit': True,
        'edit_student': student,
    })


@login_required
def student_registration_view(request):
    if not _is_admin(request.user):
        raise PermissionDenied
    if request.method == 'POST':
        student_form = StudentRegistrationForm(request.POST, request.FILES)
        father_form = ParentForm(request.POST, prefix='father')
        mother_form = ParentForm(request.POST, prefix='mother')

        if student_form.is_valid() and father_form.is_valid() and mother_form.is_valid():
            student_instance = student_form.save(commit=False)
            
            # Save parent forms if they have data
            father_obj = None
            mother_obj = None
            
            # Check if father form has any data
            father_has_data = any([
                father_form.cleaned_data.get('name'),
                father_form.cleaned_data.get('telephone_number')
            ])
            
            if father_has_data:
                father_obj = father_form.save()
            
            # Check if mother form has any data
            mother_has_data = any([
                mother_form.cleaned_data.get('name'),
                mother_form.cleaned_data.get('telephone_number')
            ])
            
            if mother_has_data:
                mother_obj = mother_form.save()
            
            # Link them directly to the student row
            student_instance.father = father_obj
            student_instance.mother = mother_obj
            student_instance.save()
            return redirect('enroll_student', student_id=student_instance.id)
    else:
        student_form = StudentRegistrationForm()
        father_form = ParentForm(prefix='father')
        mother_form = ParentForm(prefix='mother')

    return render(request, 'sis/student_registration.html', {
        'student_form': student_form,
        'father_form': father_form,
        'mother_form': mother_form,
    })

def _notify_form_teacher(request, classroom, subject, created):
    form_teacher = getattr(classroom, 'form_teacher', None)
    if not form_teacher or not form_teacher.user:
        return
    if form_teacher.user == request.user:
        return
    Notification.objects.create(
        recipient=form_teacher.user,
        title="Assessment Scores Updated",
        message=(
            f"Scores for {subject.subject_name} have been "
            f"{'added' if created else 'updated'} for {classroom.class_name}."
        ),
        notification_type='ASSESSMENT_UPDATE',
    )


# bulk score processing view
@login_required
def bulk_grade_entry_view(request, class_id, subject_id):
    if not _is_staff_or_admin(request.user):
        raise PermissionDenied
    classroom = ClassRoom.objects.get(pk=class_id)
    staff = getattr(request.user, 'staff_profile', None)

    if not request.user.is_superuser:
        has_assignment = StaffClassSubject.objects.filter(staff=staff, classroom=classroom, subject_id=subject_id).exists()
        if not has_assignment:
            messages.error(request, 'You can only enter grades for subjects and classes assigned to you.')
            return redirect('dashboard')

    students = Student.objects.filter(enrollments__classroom=classroom).distinct()
    if request.user.is_superuser:
        subjects = Subject.objects.all().order_by('subject_name')
    else:
        assigned_subject_ids = StaffClassSubject.objects.filter(staff=staff, classroom=classroom).values_list('subject_id', flat=True).distinct()
        subjects = Subject.objects.filter(id__in=assigned_subject_ids).order_by('subject_name')

    if subject_id:
        subject = Subject.objects.filter(pk=subject_id).first()
    else:
        subject = subjects.first()

    if subject is None:
        return render(request, 'sis/bulk_grade_entry.html', {
            'classroom': classroom,
            'subjects': subjects,
            'subject': None,
            'matrix': [],
            'message': 'No subjects are available yet.'
        })

    # static anchors for current tracking context
    current_session = AcademicSession.objects.filter(is_current=True).first()
    current_term_obj = Term.objects.filter(is_active=True).first()

    if request.method == 'POST':
        for student in students:
            # safety grab input values from the raw POST stream using unique field id
            class_val = request.POST.get(f'class_score_{student.id}')
            exam_val = request.POST.get(f'exam_score_{student.id}')
            
            if class_val and exam_val:
                SubjectAssessment.objects.update_or_create(
                    student=student,
                    subject=subject,
                    academic_session=current_session,
                    academic_term=current_term_obj,
                    defaults={
                        'class_score': class_val,
                        'exam_score': exam_val,
                    }
                )
        messages.success(request, f"Grades for {subject.subject_name} saved successfully!")
        return redirect('bulk_grade_entry', class_id=class_id, subject_id=subject_id)
    
    # Build up existing data list to repopulate inputs if scores are already entered
    student_marks_matrix = []
    for student in students:
        existing_assessment = SubjectAssessment.objects.filter(
                student=student, subject=subject, academic_session=current_session, academic_term=current_term_obj
            ).first()
        
        student_marks_matrix.append({
            'student': student,
            'existing_class_score': existing_assessment.class_score if existing_assessment else "",
            'existing_exam_score': existing_assessment.exam_score if existing_assessment else ""
        })

    context = {
        'classroom': classroom,
        'subject': subject,
        'subjects': subjects,
        'matrix': student_marks_matrix
    }
    return render(request, 'sis/bulk_grade_entry.html', context)

@login_required
def class_report_card_view(request, class_id):
    if not _is_staff_or_admin(request.user):
        raise PermissionDenied
    classroom = ClassRoom.objects.get(pk=class_id)

    staff = getattr(request.user, 'staff_profile', None)
    is_form_teacher = staff and staff.form_class == classroom
    has_full_access = request.user.is_superuser or is_form_teacher

    try:
        user_form_class = request.user.staff_profile.form_class
    except AttributeError:
        user_form_class = None

    current_subject_id = request.GET.get('subject_id')
    if current_subject_id and not current_subject_id.isdigit():
        current_subject_id = None
    current_subject_id = int(current_subject_id) if current_subject_id else None

    is_master = request.GET.get('master') == '1'

    if staff:
        assigned_ids = StaffClassSubject.objects.filter(staff=staff, classroom=classroom).values_list('subject_id', flat=True).distinct()
        assigned_subjects = Subject.objects.filter(id__in=assigned_ids) if assigned_ids else Subject.objects.none()
    else:
        assigned_subjects = Subject.objects.none()

    has_class_subject_assignment = request.user.is_superuser or assigned_subjects.exists()

    subjects_for_class = Subject.objects.filter(offered_in_classes__classroom=classroom).distinct().order_by('subject_name')

    active_subject = None
    if not is_master:
        if current_subject_id:
            active_subject = Subject.objects.filter(id=current_subject_id).first()
        elif assigned_subjects.exists():
            active_subject = assigned_subjects.first()
        if active_subject and active_subject not in assigned_subjects:
            active_subject = assigned_subjects.first() if assigned_subjects.exists() else None

    students = Student.objects.filter(enrollments__classroom=classroom).distinct()

    current_session = AcademicSession.objects.filter(is_current=True).first()
    current_term = Term.objects.filter(is_active=True).first()

    all_assessments = SubjectAssessment.objects.filter(
        academic_session=current_session, academic_term=current_term,
        student__in=students, subject__in=subjects_for_class,
    )
    if is_master or request.user.is_superuser:
        pass
    elif has_class_subject_assignment:
        all_assessments = all_assessments.filter(subject_id__in=assigned_subjects.values_list('id', flat=True))
    else:
        all_assessments = SubjectAssessment.objects.none()

    assessment_map = {}
    for a in all_assessments:
        assessment_map.setdefault(a.student_id, {})[a.subject_id] = a

    GRADE_REMARKS = get_grade_remarks_for_class(classroom)

    def get_remark(total):
        if total is None:
            return "—", "—"
        for floor, grade, label in GRADE_REMARKS:
            if total >= floor:
                return grade, label
        return GRADE_REMARKS[-1][1], GRADE_REMARKS[-1][2]

    report_data = []
    for student in students:
        subject_scores = {}
        for subj in subjects_for_class:
            a = assessment_map.get(student.id, {}).get(subj.id)
            if a:
                subject_scores[subj.id] = {
                    'class_score': float(a.class_score),
                    'exam_score': float(a.exam_score),
                    'total': float(a.total_score),
                }
            else:
                subject_scores[subj.id] = {'class_score': None, 'exam_score': None, 'total': None}

        grand_total = sum(
            s['total'] for s in subject_scores.values() if s['total'] is not None
        )

        row = {
            'student': student,
            'subject_scores': subject_scores,
            'grand_total': grand_total,
        }

        if active_subject:
            sc = subject_scores.get(active_subject.id, {})
            total = sc.get('total')
            grade, remark = get_remark(total)
            row['subject_class_score'] = sc.get('class_score')
            row['subject_exam_score'] = sc.get('exam_score')
            row['subject_total'] = total
            row['subject_grade'] = grade
            row['subject_remark'] = remark

        report_data.append(row)

    if is_master:
        report_data = sorted(report_data, key=lambda x: x['grand_total'], reverse=True)
    elif active_subject:
        report_data = sorted(report_data, key=lambda x: x['subject_total'] if x['subject_total'] is not None else -1, reverse=True)

    for index, row in enumerate(report_data):
        row['rank'] = index + 1

    for subject in subjects_for_class:
        scored = [
            row for row in report_data
            if row['subject_scores'][subject.id]['total'] is not None
        ]
        scored.sort(
            key=lambda r: r['subject_scores'][subject.id]['total'],
            reverse=True,
        )
        dense_rank = 0
        prev_total = None
        for row in scored:
            total = row['subject_scores'][subject.id]['total']
            if total != prev_total:
                dense_rank += 1
                prev_total = total
            row['subject_scores'][subject.id]['subject_position'] = dense_rank

    subject_position_map = {}
    if active_subject:
        ranked_subject = [r for r in report_data if r['subject_total'] is not None]
        for idx, r in enumerate(ranked_subject):
            subject_position_map[r['student'].id] = idx + 1

    if request.user.is_superuser:
        classrooms = ClassRoom.objects.all()
    else:
        assigned_ids = StaffClassSubject.objects.filter(staff=staff).values_list('classroom_id', flat=True).distinct()
        classrooms = ClassRoom.objects.filter(id__in=assigned_ids) if assigned_ids else ClassRoom.objects.none()
        if staff and staff.form_class and staff.form_class_id not in classrooms.values_list('id', flat=True):
            classrooms = ClassRoom.objects.filter(Q(id__in=classrooms) | Q(id=staff.form_class_id)).distinct()
    has_graded_records = bool(assessment_map)

    can_modify_grades = False
    if staff:
        if current_subject_id:
            can_modify_grades = StaffClassSubject.objects.filter(
                staff=staff, classroom=classroom, subject_id=current_subject_id
            ).exists()
        else:
            can_modify_grades = assigned_subjects.exists()

    all_subject_ids = set(subjects_for_class.values_list('id', flat=True))
    assigned_subject_ids_set = set(assigned_subjects.values_list('id', flat=True))
    can_edit_master = bool(assigned_subject_ids_set)

    verification = GradeVerification.objects.filter(
        classroom=classroom, academic_session=current_session, academic_term=current_term
    ).first()

    assessment_json = {}
    for sid, subj_map in assessment_map.items():
        assessment_json[str(sid)] = {}
        for subj_id, a in subj_map.items():
            assessment_json[str(sid)][str(subj_id)] = {
                'cs': float(a.class_score),
                'es': float(a.exam_score),
            }

    students_json = [{'id': s.id, 'name': f"{s.first_name} {s.last_name}"} for s in students]
    subjects_json = [{'id': s.id, 'name': s.subject_name} for s in subjects_for_class]

    is_admin_landing = request.user.is_superuser and not is_master and not active_subject

    return render(request, 'sis/class_report.html', {
        'classroom': classroom,
        'report_data': report_data,
        'is_admin_landing': is_admin_landing,
        'classrooms': classrooms,
        'assigned_classes': classrooms,
        'current_class_id': classroom.id,
        'user_form_class': user_form_class,
        'has_graded_records': has_graded_records,
        'is_form_teacher': is_form_teacher,
        'is_master': is_master,
        'has_full_access': has_full_access,
        'has_class_subject_assignment': has_class_subject_assignment,
        'verification': verification,
        'assigned_subjects': assigned_subjects,
        'subjects_for_class': subjects_for_class,
        'active_subject': active_subject,
        'current_subject_id': current_subject_id,
        'subject_position_map': subject_position_map,
        'can_modify_grades': can_modify_grades,
        'can_edit_master': can_edit_master,
        'assessment_json': assessment_json,
        'students_json': students_json,
        'subjects_json': subjects_json,
    })


@login_required
@require_POST
def api_check_completeness(request, class_id):
    if not _is_staff_or_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    classroom = ClassRoom.objects.get(pk=class_id)
    staff = getattr(request.user, 'staff_profile', None)
    is_form_teacher = staff and staff.form_class == classroom
    has_full_access = request.user.is_superuser or is_form_teacher

    if not has_full_access:
        return JsonResponse({'error': 'Only the form teacher or admin can generate reports'}, status=403)

    current_session = AcademicSession.objects.filter(is_current=True).first()
    current_term = Term.objects.filter(is_active=True).first()
    if not current_session or not current_term:
        return JsonResponse({'complete': True, 'missing': []})

    subjects_for_class = Subject.objects.filter(offered_in_classes__classroom=classroom).distinct()
    students = Student.objects.filter(enrollments__classroom=classroom).distinct()

    all_assessments = SubjectAssessment.objects.filter(
        academic_session=current_session, academic_term=current_term,
        student__in=students, subject__in=subjects_for_class,
    )
    assessment_map = {}
    for a in all_assessments:
        assessment_map.setdefault(a.student_id, {})[a.subject_id] = a

    missing = []
    for subject in subjects_for_class:
        teacher_name = ""
        scs = StaffClassSubject.objects.filter(classroom=classroom, subject=subject).select_related('staff').first()
        if scs:
            teacher_name = f"{scs.staff.first_name} {scs.staff.last_name}"

        missing_class = []
        missing_exam = []
        for student in students:
            a = assessment_map.get(student.id, {}).get(subject.id)
            if not a:
                missing_class.append(student.first_name)
                missing_exam.append(student.first_name)
            else:
                if a.class_score is None:
                    missing_class.append(student.first_name)
                if a.exam_score is None:
                    missing_exam.append(student.first_name)

        if missing_class or missing_exam:
            entry = {'subject': subject.subject_name, 'teacher': teacher_name}
            if missing_class:
                entry['missing_class_count'] = len(missing_class)
                entry['missing_class_students'] = missing_class[:5]
            if missing_exam:
                entry['missing_exam_count'] = len(missing_exam)
                entry['missing_exam_students'] = missing_exam[:5]
            missing.append(entry)

    return JsonResponse({'complete': len(missing) == 0, 'missing': missing})


@login_required
@require_POST
def api_edit_assessment(request):
    if not _is_staff_or_admin(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    student_id = data.get('student_id')
    subject_id = data.get('subject_id')
    class_score = data.get('class_score')
    exam_score = data.get('exam_score')

    if not student_id or not subject_id:
        return JsonResponse({'error': 'student_id and subject_id are required'}, status=400)

    try:
        class_score = float(class_score) if class_score is not None else None
        exam_score = float(exam_score) if exam_score is not None else None
    except (TypeError, ValueError):
        return JsonResponse({'error': 'Invalid score values'}, status=400)

    if class_score is not None and not (0 <= class_score <= 30):
        return JsonResponse({'error': 'Class score must be between 0 and 30'}, status=400)
    if exam_score is not None and not (0 <= exam_score <= 70):
        return JsonResponse({'error': 'Exam score must be between 0 and 70'}, status=400)

    student = get_object_or_404(Student, pk=student_id)
    subject = get_object_or_404(Subject, pk=subject_id)

    staff = getattr(request.user, 'staff_profile', None)
    if not staff:
        return JsonResponse({'error': 'You are not assigned to teach this subject in this class'}, status=403)
    enrollment = student.enrollments.order_by('-date_enrolled').first()
    if not enrollment:
        return JsonResponse({'error': 'Student is not enrolled'}, status=400)
    is_assigned = StaffClassSubject.objects.filter(
        staff=staff, classroom=enrollment.classroom, subject=subject
    ).exists()
    if not is_assigned:
        return JsonResponse({'error': 'You are not assigned to teach this subject in this class'}, status=403)

    current_session = AcademicSession.objects.filter(is_current=True).first()
    current_term = Term.objects.filter(is_active=True).first()
    if not current_session or not current_term:
        return JsonResponse({'error': 'No active academic session/term'}, status=400)

    assessment, created = SubjectAssessment.objects.update_or_create(
        student=student,
        subject=subject,
        academic_session=current_session,
        academic_term=current_term,
        defaults={
            'class_score': class_score if class_score is not None else 0,
            'exam_score': exam_score if exam_score is not None else 0,
        },
    )

    _notify_form_teacher(request, enrollment.classroom, subject, created)

    total = float(assessment.class_score or 0) + float(assessment.exam_score or 0)

    return JsonResponse({
        'success': True,
        'created': created,
        'student_id': student.id,
        'subject_id': subject.id,
        'class_score': float(assessment.class_score),
        'exam_score': float(assessment.exam_score),
        'total': total,
    })


@login_required
def generate_report_cards_view(request, class_id):
    if not _is_staff_or_admin(request.user):
        raise PermissionDenied
    classroom = ClassRoom.objects.get(pk=class_id)
    staff = getattr(request.user, 'staff_profile', None)
    is_form_teacher = staff and staff.form_class == classroom
    has_full_access = request.user.is_superuser or is_form_teacher

    if not has_full_access:
        raise PermissionDenied

    current_session = AcademicSession.objects.filter(is_current=True).first()
    current_term = Term.objects.filter(is_active=True).first()

    subjects_for_class = Subject.objects.filter(offered_in_classes__classroom=classroom).distinct().order_by('subject_name')
    students = Student.objects.filter(enrollments__classroom=classroom).distinct()

    all_assessments = SubjectAssessment.objects.filter(
        academic_session=current_session, academic_term=current_term,
        student__in=students, subject__in=subjects_for_class,
    )
    assessment_map = {}
    for a in all_assessments:
        assessment_map.setdefault(a.student_id, {})[a.subject_id] = a

    GRADE_REMARKS = get_grade_remarks_for_class(classroom)

    def get_remark(total):
        if total is None:
            return "—", "—"
        for floor, grade, label in GRADE_REMARKS:
            if total >= floor:
                return grade, label
        return GRADE_REMARKS[-1][1], GRADE_REMARKS[-1][2]

    report_data = []
    for student in students:
        subject_scores = {}
        for subj in subjects_for_class:
            a = assessment_map.get(student.id, {}).get(subj.id)
            if a:
                grade, remark = get_remark(float(a.total_score))
                subject_scores[subj.id] = {
                    'class_score': float(a.class_score),
                    'exam_score': float(a.exam_score),
                    'total': float(a.total_score),
                    'grade': grade,
                    'remark': remark,
                }
            else:
                subject_scores[subj.id] = {
                    'class_score': None, 'exam_score': None, 'total': None,
                    'grade': '—', 'remark': '—',
                }

        grand_total = sum(
            s['total'] for s in subject_scores.values() if s['total'] is not None
        )
        grade, remark = get_remark(grand_total)

        report_data.append({
            'student': student,
            'subject_scores': subject_scores,
            'grand_total': grand_total,
            'overall_grade': grade,
            'overall_remark': remark,
        })

    report_data = sorted(report_data, key=lambda x: x['grand_total'], reverse=True)
    for index, row in enumerate(report_data):
        row['rank'] = index + 1
        row['rank_ordinal'] = get_ordinal(row['rank'])

    subject_positions = {}
    for subj in subjects_for_class:
        scored = [(r['student'].id, r['subject_scores'][subj.id]['total'])
                  for r in report_data
                  if r['subject_scores'][subj.id]['total'] is not None]
        scored.sort(key=lambda x: x[1], reverse=True)
        for idx, (sid, _) in enumerate(scored):
            subject_positions.setdefault(subj.id, {})[sid] = idx + 1

    verification = GradeVerification.objects.filter(
        classroom=classroom, academic_session=current_session, academic_term=current_term
    ).first()

    parent_emails = []
    for row in report_data:
        student = row['student']
        if student.father and student.father.email:
            parent_emails.append(student.father.email)
        if student.mother and student.mother.email:
            parent_emails.append(student.mother.email)

    return render(request, 'sis/generate_report_hub.html', {
        'classroom': classroom,
        'report_data': report_data,
        'subjects_for_class': subjects_for_class,
        'subject_positions': subject_positions,
        'current_session': current_session,
        'current_term': current_term,
        'term_number': int(current_term.term_name.split()[-1]) if current_term else 1,
        'year_label': current_session.academic_year if current_session else '',
        'verification': verification,
        'parent_emails': parent_emails,
        'student_count': len(report_data),
    })


@login_required
def print_report_cards_view(request, class_id):
    if not _is_staff_or_admin(request.user):
        raise PermissionDenied
    classroom = ClassRoom.objects.get(pk=class_id)
    staff = getattr(request.user, 'staff_profile', None)
    is_form_teacher = staff and staff.form_class == classroom
    has_full_access = request.user.is_superuser or is_form_teacher
    if not has_full_access:
        raise PermissionDenied

    student_ids = request.POST.getlist('student_ids[]')
    reopening_date = request.POST.get('reopening_date') or request.GET.get('reopening_date') or ''
    if not student_ids:
        return render(request, 'sis/student_report_card_print.html', {
            'classroom': classroom,
            'selected_students': [],
            'subjects_for_class': [],
            'current_session': None,
            'current_term': None,
            'term_number': 1,
            'year_label': '',
            'reopening_date': reopening_date,
        })

    current_session = AcademicSession.objects.filter(is_current=True).first()
    current_term = Term.objects.filter(is_active=True).first() if current_session else None

    subjects_for_class = Subject.objects.filter(offered_in_classes__classroom=classroom).distinct().order_by('subject_name')
    
    # Get all students for ranking, not just selected ones
    all_students = Student.objects.filter(enrollments__classroom=classroom).distinct()
    selected_student_ids = set(int(sid) for sid in student_ids)

    all_assessments = SubjectAssessment.objects.filter(
        academic_session=current_session, academic_term=current_term,
        student__in=all_students, subject__in=subjects_for_class,
    )
    assessment_map = {}
    for a in all_assessments:
        assessment_map.setdefault(a.student_id, {})[a.subject_id] = a

    GRADE_REMARKS = get_grade_remarks_for_class(classroom)

    def get_remark(total):
        if total is None:
            return "—", "—"
        for floor, grade, label in GRADE_REMARKS:
            if total >= floor:
                return grade, label
        return GRADE_REMARKS[-1][1], GRADE_REMARKS[-1][2]

    report_data = []
    # Build report for all students, then filter to selected
    for student in all_students:
        subject_scores = {}
        for subj in subjects_for_class:
            a = assessment_map.get(student.id, {}).get(subj.id)
            if a:
                grade, remark = get_remark(float(a.total_score))
                subject_scores[subj.id] = {
                    'subject': subj,
                    'class_score': float(a.class_score),
                    'exam_score': float(a.exam_score),
                    'total': float(a.total_score),
                    'grade': grade,
                    'remark': remark,
                }
            else:
                subject_scores[subj.id] = {
                    'subject': subj,
                    'class_score': None, 'exam_score': None, 'total': None,
                    'grade': '—', 'remark': '—',
                }

        grand_total = sum(
            s['total'] for s in subject_scores.values() if s['total'] is not None
        )
        grade, remark = get_remark(grand_total)

        report_data.append({
            'student': student,
            'subject_scores': subject_scores,
            'grand_total': grand_total,
            'overall_grade': grade,
            'overall_remark': remark,
        })

    report_data = sorted(report_data, key=lambda x: x['grand_total'], reverse=True)
    # Rank all students
    for index, row in enumerate(report_data):
        row['rank'] = index + 1
        row['rank_ordinal'] = get_ordinal(row['rank'])

    subject_positions = {}
    for subj in subjects_for_class:
        scored = [(r['student'].id, r['subject_scores'][subj.id]['total'])
                  for r in report_data
                  if r['subject_scores'][subj.id]['total'] is not None]
        scored.sort(key=lambda x: x[1], reverse=True)
        for idx, (sid, _) in enumerate(scored):
            subject_positions.setdefault(subj.id, {})[sid] = idx + 1

    for row in report_data:
        for subj in subjects_for_class:
            row['subject_scores'][subj.id]['subject_position'] = subject_positions.get(subj.id, {}).get(row['student'].id)

    # Filter to only selected students for printing
    selected_students = [r for r in report_data if r['student'].id in selected_student_ids]

    class_size = all_students.count()

    form_master_name = ""
    if classroom.form_master:
        form_master_name = f"{classroom.form_master.first_name} {classroom.form_master.last_name}"

    return render(request, 'sis/student_report_card_print.html', {
        'classroom': classroom,
        'selected_students': selected_students,
        'subjects_for_class': subjects_for_class,
        'current_session': current_session,
        'current_term': current_term,
        'term_number': int(current_term.term_name.split()[-1]) if current_term else 1,
        'year_label': current_session.academic_year if current_session else '',
        'class_size': class_size,
        'form_master_name': form_master_name,
        'reopening_date': reopening_date,
    })


@login_required
def export_excel_view(request, class_id):
    if not _is_staff_or_admin(request.user):
        raise PermissionDenied
    classroom = ClassRoom.objects.get(pk=class_id)
    staff = getattr(request.user, 'staff_profile', None)
    is_form_teacher = staff and staff.form_class == classroom
    has_full_access = request.user.is_superuser or is_form_teacher
    if not has_full_access:
        raise PermissionDenied

    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from django.http import HttpResponse

    current_session = AcademicSession.objects.filter(is_current=True).first()
    current_term = Term.objects.filter(is_active=True).first()
    term_number = int(current_term.term_name.split()[-1]) if current_term else 1
    year_label = current_session.academic_year if current_session else ''

    subjects_for_class = Subject.objects.filter(offered_in_classes__classroom=classroom).distinct().order_by('subject_name')
    students = Student.objects.filter(enrollments__classroom=classroom).distinct()

    all_assessments = SubjectAssessment.objects.filter(
        academic_session=current_session, academic_term=current_term,
        student__in=students, subject__in=subjects_for_class,
    )
    assessment_map = {}
    for a in all_assessments:
        assessment_map.setdefault(a.student_id, {})[a.subject_id] = a

    GRADE_REMARKS = get_grade_remarks_for_class(classroom)

    def get_remark(total):
        if total is None:
            return "—", "—"
        for floor, grade, label in GRADE_REMARKS:
            if total >= floor:
                return grade, label
        return GRADE_REMARKS[-1][1], GRADE_REMARKS[-1][2]

    wb = Workbook()
    ws = wb.active
    ws.title = f"{classroom.class_name} Results"

    header_font = Font(bold=True, size=12)
    sub_header_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4 + len(subjects_for_class) * 3 + 2)
    ws.cell(row=1, column=1, value=f"{classroom.class_name} — Terminal Report ({year_label}, Term {term_number})").font = Font(bold=True, size=14)

    row_num = 3
    headers = ["Rank", "Student Name", "Admission No."]
    for subj in subjects_for_class:
        headers.extend([f"{subj.subject_name} (30%)", f"{subj.subject_name} (70%)", f"{subj.subject_name} Total"])
    headers.extend(["Grand Total", "Grade", "Remark"])

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.font = sub_header_font
        cell.border = border
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    report_data = []
    for student in students:
        subject_scores = {}
        for subj in subjects_for_class:
            a = assessment_map.get(student.id, {}).get(subj.id)
            if a:
                grade, remark = get_remark(float(a.total_score))
                subject_scores[subj.id] = {
                    'class_score': float(a.class_score),
                    'exam_score': float(a.exam_score),
                    'total': float(a.total_score),
                    'grade': grade, 'remark': remark,
                }
            else:
                subject_scores[subj.id] = {
                    'class_score': None, 'exam_score': None, 'total': None,
                    'grade': '—', 'remark': '—',
                }
        grand_total = sum(s['total'] for s in subject_scores.values() if s['total'] is not None)
        grade, remark = get_remark(grand_total)
        report_data.append({
            'student': student, 'subject_scores': subject_scores,
            'grand_total': grand_total, 'grade': grade, 'remark': remark,
        })

    report_data = sorted(report_data, key=lambda x: x['grand_total'], reverse=True)

    for idx, row in enumerate(report_data):
        r = row_num + 1 + idx
        ws.cell(row=r, column=1, value=idx + 1).border = border
        ws.cell(row=r, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=2, value=f"{row['student'].first_name} {row['student'].last_name}").border = border
        ws.cell(row=r, column=3, value=row['student'].admission_number).border = border

        col = 4
        for subj in subjects_for_class:
            sc = row['subject_scores'][subj.id]
            ws.cell(row=r, column=col, value=sc['class_score']).border = border
            ws.cell(row=r, column=col).alignment = Alignment(horizontal='center')
            ws.cell(row=r, column=col + 1, value=sc['exam_score']).border = border
            ws.cell(row=r, column=col + 1).alignment = Alignment(horizontal='center')
            ws.cell(row=r, column=col + 2, value=sc['total']).border = border
            ws.cell(row=r, column=col + 2).alignment = Alignment(horizontal='center')
            col += 3

        total_col = col
        ws.cell(row=r, column=total_col, value=row['grand_total']).border = border
        ws.cell(row=r, column=total_col).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=total_col).font = Font(bold=True)
        ws.cell(row=r, column=total_col + 1, value=row['grade']).border = border
        ws.cell(row=r, column=total_col + 1).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=total_col + 2, value=row['remark']).border = border
        ws.cell(row=r, column=total_col + 2).alignment = Alignment(horizontal='center')

    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            if not hasattr(cell, 'column_letter'):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_length + 4, 30)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"{classroom.class_name.replace(' ', '_')}_results_{year_label}_term{term_number}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def register_staff_view(request):
    departments = Department.objects.all()
    designations = Designation.objects.all()
    all_classrooms = ClassRoom.objects.all()

    if request.method == 'POST':
        form = StaffRegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            # Create the staff profile first
            staff_profile = form.save(commit=False)

            # Handle department FK
            dept_name = form.cleaned_data.get('department')
            if dept_name:
                dept, _ = Department.objects.get_or_create(name=dept_name)
                staff_profile.department = dept

            # Handle designation FK
            desig_name = form.cleaned_data.get('designation')
            if desig_name:
                desig, _ = Designation.objects.get_or_create(name=desig_name)
                staff_profile.designation = desig

            # Parse full_name into first/last/other if the individual fields weren't
            # populated directly by the form (the form still carries `full_name`).
            full = getattr(staff_profile, 'full_name', '') or ''
            if full and not staff_profile.first_name:
                parts = full.strip().split(None, 2)
                staff_profile.first_name = parts[0] if len(parts) > 0 else ''
                staff_profile.last_name = parts[-1] if len(parts) > 1 else ''
                staff_profile.other_names = ' '.join(parts[1:-1]) if len(parts) > 2 else ''

            # Create a corresponding user account using staff_id as username
            username = staff_profile.staff_id
            password = 'staff123'  # Default password

            # Create user if it doesn't exist
            user, created = User.objects.get_or_create(
                username=username,
                defaults={
                    'first_name': staff_profile.first_name,
                    'last_name': staff_profile.last_name,
                    'email': staff_profile.email,
                    'is_staff': True,
                }
            )

            # Set the password if user was just created
            if created:
                user.set_password(password)
                user.save()

            # Link the user to the staff profile
            staff_profile.user = user
            staff_profile.save()
            form.save_m2m()

            # Parse assignments_json and bulk-create StaffClassSubject records
            assignments_raw = request.POST.get('assignments_json', '')
            if assignments_raw:
                try:
                    import json
                    assignments = json.loads(assignments_raw)
                    for class_id_str, subject_ids in assignments.items():
                        for subj_id in subject_ids:
                            StaffClassSubject.objects.get_or_create(
                                staff=staff_profile,
                                classroom_id=int(class_id_str),
                                subject_id=int(subj_id)
                            )
                except (json.JSONDecodeError, ValueError):
                    pass

            messages.success(request, f'Staff member registered successfully. Username: {username}, Password: {password}')
            return redirect('staff_list')
    else:
        form = StaffRegistrationForm()

    return render(request, 'sis/register_staff.html', {
        'form': form,
        'departments': departments,
        'designations': designations,
        'all_classrooms': all_classrooms,
    })


@login_required
def staff_list_view(request):
    queryset = StaffProfile.objects.select_related('user', 'designation', 'department', 'form_class').all()

    designation = request.GET.get('designation', '').strip()
    department_id = request.GET.get('department', '').strip()
    form_class_id = request.GET.get('form_class', '').strip()

    if designation and designation != 'all':
        queryset = queryset.filter(designation__name__iexact=designation)
    if department_id and department_id != 'all':
        queryset = queryset.filter(department_id=department_id)
    if form_class_id and form_class_id != 'all':
        if form_class_id == 'none':
            queryset = queryset.filter(form_class__isnull=True)
        else:
            queryset = queryset.filter(form_class_id=form_class_id)

    designations = Designation.objects.all()
    departments = Department.objects.all()
    classes = ClassRoom.objects.all()

    return render(request, 'sis/staff_list.html', {
        'staff_members': queryset,
        'designations': designations,
        'departments': departments,
        'classes': classes,
        'selected_designation': designation or 'all',
        'selected_department': department_id or 'all',
        'selected_form_class': form_class_id or 'all',
        'total_count': queryset.count(),
        'can_export': request.user.is_superuser,
    })


@login_required
def export_staff_excel(request):
    from datetime import datetime
    import io
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    if not _is_admin(request.user):
        raise PermissionDenied

    staff_members = StaffProfile.objects.select_related(
        'user', 'designation', 'department', 'form_class'
    ).prefetch_related('subject_areas').order_by('last_name', 'first_name')

    columns = [
        "Staff ID", "Title", "First Name", "Other Names", "Last Name", "Gender",
        "Date of Birth", "Designation", "Department", "Form Class", "Email",
        "Phone Number", "Employment Type", "Date of Appointment",
        "Year of Last Promotion", "Qualification", "Certificate",
        "Institution Completed", "Year Completed", "Subject Areas",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Staff Directory"

    header_font = Font(bold=True, size=14)
    sub_header_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.cell(row=1, column=1, value="STAFF DIRECTORY").font = header_font

    for col_idx, name in enumerate(columns, 1):
        cell = ws.cell(row=3, column=col_idx, value=name)
        cell.font = sub_header_font
        cell.border = border
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    row = 4
    for s in staff_members:
        subject_areas = ", ".join(
            sub.subject_name for sub in s.subject_areas.all()
        )
        values = [
            s.staff_id, s.title, s.first_name, s.other_names or '', s.last_name, s.gender,
            s.dob.strftime('%Y-%m-%d') if s.dob else '',
            s.designation.name if s.designation else '',
            s.department.name if s.department else '',
            s.form_class.class_name if s.form_class else '',
            s.email, s.phone_number or '', s.employment_type,
            s.date_of_appointment.strftime('%Y-%m-%d') if s.date_of_appointment else '',
            s.year_of_last_promotion if s.year_of_last_promotion is not None else '',
            s.qualification, s.certificate,
            s.name_of_institution_completed, s.year_completed,
            subject_areas,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = border
        row += 1

    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            if not hasattr(cell, 'column_letter'):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_length + 4, 30)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Staff_Directory_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_students_excel(request):
    from datetime import datetime
    import io
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    user = request.user
    if _is_admin(user):
        students = Student.objects.filter(is_alumni=False).select_related(
            'classroom', 'father', 'mother'
        ).order_by('last_name', 'first_name')
    elif hasattr(user, 'staff_profile'):
        staff = user.staff_profile
        visible_class_ids = set(StaffClassSubject.objects.filter(
            staff=staff
        ).values_list('classroom_id', flat=True))
        if staff.form_class:
            visible_class_ids.add(staff.form_class_id)
        visible_class_ids.update(
            ClassRoom.objects.filter(form_master=staff).values_list('id', flat=True)
        )
        students = Student.objects.filter(
            classroom_id__in=visible_class_ids, is_alumni=False
        ).select_related('classroom', 'father', 'mother').order_by('last_name', 'first_name')
    else:
        raise PermissionDenied

    columns = [
        "Admission No.", "First Name", "Other Names", "Last Name", "Gender",
        "Date of Birth", "Class", "Status",
        "Father Name", "Father Phone", "Father Email",
        "Mother Name", "Mother Phone", "Mother Email",
        "Parent Address",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Students Directory"

    header_font = Font(bold=True, size=14)
    sub_header_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.cell(row=1, column=1, value="STUDENTS DIRECTORY").font = header_font

    for col_idx, name in enumerate(columns, 1):
        cell = ws.cell(row=3, column=col_idx, value=name)
        cell.font = sub_header_font
        cell.border = border
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    row = 4
    for s in students:
        father = s.father
        mother = s.mother
        address = ''
        if father and father.residential_address:
            address = father.residential_address
        elif mother and mother.residential_address:
            address = mother.residential_address

        values = [
            s.admission_number, s.first_name, s.other_names or '', s.last_name, s.gender,
            s.dob.strftime('%Y-%m-%d'), s.classroom.class_name if s.classroom else '', s.status,
            father.name if father else '', father.telephone_number if father else '', father.email if father else '',
            mother.name if mother else '', mother.telephone_number if mother else '', mother.email if mother else '',
            address,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = border
        row += 1

    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            if not hasattr(cell, 'column_letter'):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_length + 4, 30)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Students_Directory_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_alumni_excel(request):
    from datetime import datetime
    import io
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    user = request.user
    if not _is_admin(user) and not hasattr(user, 'staff_profile'):
        raise PermissionDenied

    students = Student.objects.filter(is_alumni=True).select_related(
        'classroom', 'father', 'mother'
    ).order_by('last_name', 'first_name')

    columns = [
        "First Name", "Other Names", "Last Name", "Admission No.",
        "Gender", "Last Class", "Status",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Alumni Directory"

    header_font = Font(bold=True, size=14)
    sub_header_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.cell(row=1, column=1, value="ALUMNI DIRECTORY").font = header_font

    for col_idx, name in enumerate(columns, 1):
        cell = ws.cell(row=3, column=col_idx, value=name)
        cell.font = sub_header_font
        cell.border = border
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    row = 4
    for s in students:
        values = [
            s.first_name, s.other_names or '', s.last_name, s.admission_number,
            s.gender,
            s.classroom.class_name if s.classroom else '',
            s.status,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = border
        row += 1

    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            if not hasattr(cell, 'column_letter'):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_length + 4, 30)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Alumni_Directory_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_parents_excel(request):
    from datetime import datetime
    import io
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    user = request.user

    children_prefetch = Prefetch(
        'father_of',
        queryset=Student.objects.select_related('classroom').filter(is_alumni=False),
        to_attr='father_children'
    )
    mother_prefetch = Prefetch(
        'mother_of',
        queryset=Student.objects.select_related('classroom').filter(is_alumni=False),
        to_attr='mother_children'
    )

    if user.is_superuser:
        parents = Parent.objects.prefetch_related(children_prefetch, mother_prefetch).order_by('name')
    elif hasattr(user, 'staff_profile'):
        staff = user.staff_profile
        classroom_ids = StaffClassSubject.objects.filter(
            staff=staff
        ).values_list('classroom_id', flat=True).distinct()
        parents = Parent.objects.filter(
            Q(father_of__classroom_id__in=classroom_ids) |
            Q(mother_of__classroom_id__in=classroom_ids)
        ).distinct().prefetch_related(children_prefetch, mother_prefetch).order_by('name')
    else:
        raise PermissionDenied

    columns = [
        "Parent Name", "Phone", "Email", "Occupation", "Residential Address",
        "No. of Children", "Children (Name - Class)",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Parents Directory"

    header_font = Font(bold=True, size=14)
    sub_header_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    ws.cell(row=1, column=1, value="PARENTS DIRECTORY").font = header_font

    for col_idx, name in enumerate(columns, 1):
        cell = ws.cell(row=3, column=col_idx, value=name)
        cell.font = sub_header_font
        cell.border = border
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

    row = 4
    for parent in parents:
        father_children = list(getattr(parent, 'father_children', []))
        mother_children = list(getattr(parent, 'mother_children', []))
        children = father_children + mother_children
        children_str = "; ".join(
            f"{child.first_name} {child.last_name} - {child.classroom.class_name if child.classroom else 'N/A'}"
            for child in children
        )
        values = [
            parent.name or '', parent.telephone_number or '', parent.email or '',
            parent.occupation or '', parent.residential_address or '',
            len(children), children_str,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = border
        row += 1

    for col in ws.columns:
        max_length = 0
        col_letter = None
        for cell in col:
            if not hasattr(cell, 'column_letter'):
                continue
            if col_letter is None:
                col_letter = cell.column_letter
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Parents_Directory_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def staff_detail_view(request, staff_id):
    staff_member = get_object_or_404(StaffProfile, pk=staff_id)
    all_classes = ClassRoom.objects.all()
    assignments = StaffClassSubject.objects.filter(
        staff=staff_member
    ).select_related('classroom', 'subject')
    return render(request, 'sis/staff_detail.html', {
        'staff_member': staff_member,
        'all_classes': all_classes,
        'assignments': assignments,
    })


@login_required
@require_POST
def assign_form_class(request, staff_id):
    if not (request.user.is_superuser or request.user.is_staff):
        return JsonResponse({'success': False, 'error': 'Unauthorized'}, status=403)

    data = json.loads(request.body)
    class_id = data.get('class_id')

    target_staff = get_object_or_404(StaffProfile, pk=staff_id)
    target_class = get_object_or_404(ClassRoom, pk=class_id)

    try:
        old_teacher = target_class.form_teacher
        if old_teacher:
            old_teacher.form_class = None
            old_teacher.save()
    except StaffProfile.DoesNotExist:
        pass

    target_staff.form_class = target_class
    target_staff.save()

    return JsonResponse({'success': True})


@login_required
def staff_edit_view(request, staff_id):
    import json
    staff_member = get_object_or_404(StaffProfile, pk=staff_id)
    departments = Department.objects.all()
    designations = Designation.objects.all()
    all_classrooms = ClassRoom.objects.all()

    # Build existing assignments for the JS picker
    existing_assignments = {}
    for scs in staff_member.assigned_classes_subjects.select_related('classroom', 'subject').all():
        class_id = str(scs.classroom_id)
        if class_id not in existing_assignments:
            existing_assignments[class_id] = []
        existing_assignments[class_id].append(str(scs.subject_id))

    if request.method == 'POST':
        form = StaffRegistrationForm(request.POST, request.FILES, instance=staff_member)
        if form.is_valid():
            staff_profile = form.save(commit=False)

            dept_name = form.cleaned_data.get('department')
            if dept_name:
                dept, _ = Department.objects.get_or_create(name=dept_name)
                staff_profile.department = dept
            else:
                staff_profile.department = None

            desig_name = form.cleaned_data.get('designation')
            if desig_name:
                desig, _ = Designation.objects.get_or_create(name=desig_name)
                staff_profile.designation = desig
            else:
                staff_profile.designation = None

            user = staff_profile.user
            if user:
                user.first_name = staff_profile.first_name
                user.last_name = staff_profile.last_name
                user.email = staff_profile.email
                user.save()

            staff_profile.save()
            form.save_m2m()

            # Replace all ClassSubject assignments
            staff_profile.assigned_classes_subjects.all().delete()
            assignments_raw = request.POST.get('assignments_json', '')
            if assignments_raw:
                try:
                    assignments = json.loads(assignments_raw)
                    for class_id_str, subject_ids in assignments.items():
                        for subj_id in subject_ids:
                            StaffClassSubject.objects.get_or_create(
                                staff=staff_profile,
                                classroom_id=int(class_id_str),
                                subject_id=int(subj_id)
                            )
                except (json.JSONDecodeError, ValueError):
                    pass

            messages.success(request, 'Staff member updated successfully.')
            return redirect('staff_detail', staff_id=staff_profile.id)
    else:
        form = StaffRegistrationForm(instance=staff_member)
        if staff_member.department:
            form.fields['department'].initial = staff_member.department.name
        if staff_member.designation:
            form.fields['designation'].initial = staff_member.designation.name

    return render(request, 'sis/register_staff.html', {
        'form': form,
        'departments': departments,
        'designations': designations,
        'all_classrooms': all_classrooms,
        'is_edit': True,
        'edit_staff': staff_member,
        'existing_assignments_json': json.dumps(existing_assignments),
    })


@login_required
def enroll_student_view(request, student_id):
    student = Student.objects.filter(pk=student_id).first()
    if not student:
        messages.error(request, 'Student not found.')
        return redirect('student_list')

    current_session = AcademicSession.objects.filter(is_current=True).first()

    if request.method == 'POST':
        form = EnrollmentForm(request.POST, session=current_session)
        if form.is_valid():
            enrollment = form.save(commit=False)
            enrollment.student = student
            enrollment.save()
            student.classroom = enrollment.classroom
            student.save(update_fields=['classroom'])
            messages.success(request, f"Student {student.first_name} {student.other_names} {student.last_name} successfully enrolled in {enrollment.classroom}!")
            return redirect('student_list')
    else:
        form = EnrollmentForm(session=current_session)

    subjects = Subject.objects.all().order_by('subject_name')

    return render(request, 'sis/enroll_student.html', {
        'student': student,
        'form': form,
        'subjects': subjects,
        'classrooms': ClassRoom.objects.all(),
    })


def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')

    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            messages.success(request, f"Welcome back, {user.username}!")

            if user.is_superuser:
                return redirect('dashboard')

            if user.is_staff or hasattr(user, 'staff_profile'):
                return redirect('dashboard')

            return redirect('dashboard')
        else:
            messages.error(request, 'Invalid credentials. Please try again.')
    else:
        form = AuthenticationForm()

    return render(request, 'sis/login.html', {'form': form})


def logout_view(request):
    logout(request)
    return redirect('login')


@login_required
def class_enrollment_portal_view(request):
    if not _is_staff_or_admin(request.user):
        raise PermissionDenied

    current_session = AcademicSession.objects.filter(is_current=True).first()
    current_term = Term.objects.filter(is_active=True).first() if current_session else None

    is_promotional_term = current_term and current_term.term_name == 'Term 3'

    if not is_promotional_term:
        term_label = current_term.term_name if current_term else "no active term"
        session_label = current_session.academic_year if current_session else "N/A"
        messages.warning(
            request,
            f"Promotion Portal is locked during {term_label}. "
            f"Class promotions can only be executed at the end of Term 3 "
            f"(Current Session: {session_label} — {term_label})."
        )
        return render(request, 'sis/promotion_locked.html', {
            'current_session': current_session,
            'current_term': current_term,
        })

    staff = getattr(request.user, 'staff_profile', None)
    if request.user.is_superuser:
        classrooms = ClassRoom.objects.all()
    else:
        if staff and staff.form_class:
            classrooms = ClassRoom.objects.filter(id=staff.form_class.id)
        else:
            classrooms = ClassRoom.objects.none()

    source_class = None
    students_data = []
    promotion_criteria = None
    search_query = request.GET.get('search_query', '').strip()
    source_class_id = request.GET.get('source_class_id')

    target_classes = []
    default_target_id = None
    is_graduated = False

    if source_class_id:
        source_class = get_object_or_404(ClassRoom, pk=source_class_id)

        next_cls = source_class.get_next_class()
        higher_classes = list(source_class.get_higher_classes())

        if next_cls:
            target_classes = [next_cls] + [c for c in higher_classes if c.id != next_cls.id]
            default_target_id = next_cls.id
        elif higher_classes:
            target_classes = higher_classes
            default_target_id = higher_classes[0].id
        else:
            is_graduated = True

        criteria_qs = PromotionCriteria.objects.filter(classroom=source_class)
        promotion_criteria = criteria_qs.first()
        min_score = float(promotion_criteria.min_grand_total) if promotion_criteria else 50.00

        enrolled_ids = Enrollment.objects.filter(
            classroom=source_class, academic_session=current_session, academic_term=current_term
        ).values_list('student_id', flat=True)

        students = Student.objects.filter(pk__in=enrolled_ids)

        if search_query:
            students = students.filter(
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(admission_number__icontains=search_query)
            )

        class_subject_names = list(
            ClassSubject.objects.filter(classroom=source_class)
            .values_list('subject__subject_name', flat=True)
        )

        for student in students:
            assessments = SubjectAssessment.objects.filter(
                student=student, academic_session=current_session, academic_term=current_term
            )
            grand_total = sum(a.total_score for a in assessments)
            eligible = grand_total >= min_score

            students_data.append({
                'student': student,
                'grand_total': grand_total,
                'eligible': eligible,
                'subjects': class_subject_names,
            })

        students_data.sort(key=lambda x: x['grand_total'], reverse=True)

    if request.method == 'POST':
        selected_ids = request.POST.getlist('selected_students')
        next_class_id = request.POST.get('next_class_id')
        src_id = request.POST.get('source_class_id')

        if next_class_id and src_id:
            next_class = get_object_or_404(ClassRoom, pk=next_class_id)
            src_class = get_object_or_404(ClassRoom, pk=src_id)

            source_ids = set(
                Enrollment.objects.filter(
                    classroom=src_class,
                ).values_list('student_id', flat=True)
            )
            selected_set = set(int(sid) for sid in selected_ids)
            held_back_ids = source_ids - selected_set

            promoted_count = 0
            held_count = 0

            for sid in selected_set:
                student = get_object_or_404(Student, pk=sid)
                student.pending_next_class = next_class
                student.promotion_status = 'APPROVED'
                student.save(update_fields=['pending_next_class', 'promotion_status'])
                promoted_count += 1

            for sid in held_back_ids:
                student = get_object_or_404(Student, pk=sid)
                student.pending_next_class = None
                student.promotion_status = 'HELD_BACK'
                student.save(update_fields=['pending_next_class', 'promotion_status'])
                held_count += 1

            from django.contrib.auth.models import User as DjangoUser
            admin_users = DjangoUser.objects.filter(is_superuser=True)
            for admin_user in admin_users:
                Notification.objects.create(
                    recipient=admin_user,
                    title="Student Promotions Processed",
                    message=(
                        f"Promotion approvals for {src_class.class_name} recorded: "
                        f"{promoted_count} approved for {next_class.class_name}, "
                        f"{held_count} held back."
                    ),
                    notification_type='PROMOTION',
                )

            messages.success(
                request,
                f"Promotion approvals recorded: {promoted_count} student(s) approved for {next_class.class_name}, "
                f"{held_count} held back. Changes will take effect at session rollover."
            )
            return redirect('class_enrollment_portal')

    context = {
        'classrooms': classrooms,
        'current_session': current_session,
        'current_term': current_term,
        'source_class': source_class,
        'promotion_criteria': promotion_criteria,
        'students_data': students_data,
        'search_query': search_query,
        'target_classes': target_classes,
        'default_target_id': default_target_id,
        'is_graduated': is_graduated,
    }
    return render(request, 'sis/class_enrollment_portal.html', context)


@login_required
def configure_session_view(request):
    if not _is_admin(request.user):
        raise PermissionDenied

    sessions = AcademicSession.objects.all()
    terms = Term.objects.all()

    if request.method == 'POST':
        if 'create_new_session' in request.POST:
            year_string = request.POST.get('new_academic_year', '').strip()
            term_name = request.POST.get('new_term_name', '').strip()

            if year_string and term_name:
                with transaction.atomic():
                    session, created = AcademicSession.objects.get_or_create(
                        academic_year=year_string
                    )
                    AcademicSession.objects.update(is_current=False)
                    session.is_current = True
                    session.save()

                    term_obj, term_created = Term.objects.get_or_create(
                        session=session, term_name=term_name
                    )
                    Term.objects.update(is_active=False)
                    term_obj.is_active = True
                    term_obj.save()

                if created:
                    messages.success(request, f"Academic session '{year_string}' created and activated with {term_name}.")
                else:
                    messages.success(request, f"Session '{year_string}' already existed — it has been re-activated as the current session with {term_name} as the active term.")
            else:
                messages.error(request, "Both academic year and term are required.")

            return redirect('configure_session')

        session_id = request.POST.get('academic_session')
        term_id = request.POST.get('term')

        if not session_id:
            messages.error(request, "Please select an Academic Session.")
            return redirect('configure_session')

        selected_session = AcademicSession.objects.filter(pk=session_id).first()
        if not selected_session:
            messages.error(request, "Invalid session selected.")
            return redirect('configure_session')

        if not term_id:
            messages.error(
                request,
                "Please select an Active Term. The active term must belong to the selected session "
                "so reports and records stay consistent; if this session has no terms yet, "
                "use 'Create New Academic Session' to add one."
            )
            return redirect('configure_session')

        selected_term = Term.objects.filter(pk=term_id).first()
        if not selected_term:
            messages.error(request, "Invalid term selected.")
            return redirect('configure_session')

        if selected_term.session_id != selected_session.pk:
            messages.error(
                request,
                f"'{selected_term.term_name}' belongs to {selected_term.session.academic_year}, "
                f"not {selected_session.academic_year}. Please select a term from the correct session."
            )
            return redirect('configure_session')

        with transaction.atomic():
            AcademicSession.objects.update(is_current=False)
            selected_session.is_current = True
            selected_session.save()

            Term.objects.update(is_active=False)
            selected_term.is_active = True
            selected_term.save()

        messages.success(request, "Academic environment successfully updated!")
        return redirect('configure_session')

    context = {
        'sessions': sessions,
        'terms': terms,
        'current_session': AcademicSession.objects.filter(is_current=True).first(),
        'current_term': Term.objects.filter(is_active=True).first(),
    }
    return render(request, 'sis/configure_session.html', context)


@login_required
def academic_year_rollover_view(request):
    if not _is_admin(request.user):
        raise PermissionDenied

    current_session = AcademicSession.objects.filter(is_current=True).first()
    current_term = Term.objects.filter(is_active=True).first() if current_session else None

    approved_students = Student.objects.filter(
        promotion_status='APPROVED', pending_next_class__isnull=False
    ).select_related('classroom', 'pending_next_class')

    held_back_students = Student.objects.filter(promotion_status='HELD_BACK')

    graduated_students = Student.objects.filter(
        promotion_status='APPROVED', pending_next_class__isnull=True, classroom__next_class__isnull=True
    )

    if request.method == 'POST':
        confirm = request.POST.get('confirm_rollover')
        target_session_id = request.POST.get('target_session_id')

        if confirm != 'YES_EXECUTE_ROLLOVER':
            messages.error(request, "Confirmation text did not match. Type YES_EXECUTE_ROLLOVER to proceed.")
            return redirect('academic_year_rollover')

        if not target_session_id:
            messages.error(request, "Please select a Target Academic Session for promoted students.")
            return redirect('academic_year_rollover')

        target_session = AcademicSession.objects.filter(pk=target_session_id).first()
        if not target_session:
            messages.error(request, "Invalid target session selected.")
            return redirect('academic_year_rollover')

        if current_session and target_session.pk == current_session.pk:
            messages.error(request, "Target session cannot be the same as the current session.")
            return redirect('academic_year_rollover')

        target_term, _ = Term.objects.get_or_create(session=target_session, term_name='Term 1')

        with transaction.atomic():
            graduated_count = Student.objects.filter(
                promotion_status='APPROVED',
                pending_next_class__isnull=True,
                classroom__next_class__isnull=True,
            ).update(is_alumni=True, promotion_status='NEUTRAL')

            promoted_count = 0
            for student in Student.objects.filter(
                promotion_status='APPROVED', pending_next_class__isnull=False
            ).select_related('pending_next_class'):
                new_class = student.pending_next_class

                student.classroom = new_class
                student.pending_next_class = None
                student.promotion_status = 'NEUTRAL'
                student.save(update_fields=['classroom', 'pending_next_class', 'promotion_status'])

                Enrollment.objects.update_or_create(
                    student=student,
                    classroom=new_class,
                    academic_session=target_session,
                    academic_term=target_term,
                )
                promoted_count += 1

            held_back_students_qs = Student.objects.filter(promotion_status='HELD_BACK')
            for student in held_back_students_qs:
                if student.classroom:
                    Enrollment.objects.update_or_create(
                        student=student,
                        classroom=student.classroom,
                        academic_session=target_session,
                        academic_term=target_term,
                    )
            held_count = held_back_students_qs.update(promotion_status='NEUTRAL')

        messages.success(
            request,
            f"Academic year rollover complete: {promoted_count} promoted, "
            f"{graduated_count} graduated to alumni, {held_count} held back. "
            f"All students enrolled in {target_session.academic_year} Term 1."
        )
        return redirect('configure_session')

    available_sessions = AcademicSession.objects.exclude(pk=current_session.pk).order_by('academic_year') if current_session else AcademicSession.objects.all().order_by('academic_year')

    context = {
        'current_session': current_session,
        'current_term': current_term,
        'approved_students': approved_students,
        'held_back_students': held_back_students,
        'graduated_students': graduated_students,
        'total_pending': approved_students.count() + held_back_students.count(),
        'available_sessions': available_sessions,
    }
    return render(request, 'sis/academic_year_rollover.html', context)


@login_required
def global_search_view(request):
    q = request.GET.get('q', '').strip()
    scope = request.GET.get('scope', 'dashboard')
    if len(q) < 2:
        return JsonResponse({'results': []})

    full_name_q = Q(first_name__icontains=q) | Q(last_name__icontains=q)
    results = []

    if scope == 'alumni':
        for s in Student.objects.filter(full_name_q | Q(admission_number__icontains=q), is_alumni=True)[:8]:
            results.append({
                'id': s.id,
                'name': f"{s.first_name} {s.last_name}",
                'extra': s.admission_number,
                'type': 'Student',
                'url': reverse('student_detail', args=[s.id]),
            })

    elif scope == 'students':
        for s in Student.objects.filter(full_name_q | Q(admission_number__icontains=q)).exclude(is_alumni=True)[:8]:
            results.append({
                'id': s.id,
                'name': f"{s.first_name} {s.last_name}",
                'extra': s.admission_number,
                'type': 'Student',
                'url': reverse('student_detail', args=[s.id]),
            })

    elif scope == 'staff':
        for st in StaffProfile.objects.filter(
            Q(first_name__icontains=q) | Q(last_name__icontains=q) | Q(staff_id__icontains=q)
        )[:8]:
            results.append({
                'id': st.id,
                'name': f"{st.first_name} {st.last_name}",
                'extra': st.staff_id,
                'type': 'Staff',
                'url': reverse('staff_detail', args=[st.id]),
            })

    elif scope == 'parents':
        for p in Parent.objects.filter(
            Q(name__icontains=q) |
            Q(email__icontains=q) |
            Q(telephone_number__icontains=q) |
            Q(father_of__admission_number__icontains=q) |
            Q(father_of__first_name__icontains=q) |
            Q(father_of__last_name__icontains=q) |
            Q(father_of__other_names__icontains=q) |
            Q(mother_of__admission_number__icontains=q) |
            Q(mother_of__first_name__icontains=q) |
            Q(mother_of__last_name__icontains=q) |
            Q(mother_of__other_names__icontains=q)
        ).distinct()[:8]:
            results.append({
                'id': p.id,
                'name': p.name or 'Unnamed',
                'extra': p.email or '',
                'type': 'Parent',
                'url': reverse('parent_detail', args=[p.id]),
            })

    elif scope == 'classes':
        for c in ClassRoom.objects.filter(class_name__icontains=q)[:5]:
            results.append({
                'id': c.id,
                'name': c.class_name,
                'extra': 'Class',
                'type': 'Class',
                'url': '#',
                'action': f'openClassModal({c.id})',
            })
        for s in Subject.objects.filter(subject_name__icontains=q)[:5]:
            results.append({
                'id': s.id,
                'name': s.subject_name,
                'extra': 'Subject',
                'type': 'Subject',
                'url': '#',
                'action': f'openSubjectModal({s.id})',
            })

    elif scope == 'timetables':
        for slot in TimetableSlot.objects.filter(
            Q(timetable__student_class__class_name__icontains=q) |
            Q(subject__subject_name__icontains=q) |
            Q(teacher__first_name__icontains=q) |
            Q(teacher__last_name__icontains=q)
        ).select_related('timetable__student_class', 'subject', 'teacher').distinct()[:8]:
            teacher_name = f"{slot.teacher.first_name} {slot.teacher.last_name}" if slot.teacher else 'Unassigned'
            results.append({
                'id': slot.id,
                'name': f"{slot.timetable.student_class.class_name} — {slot.subject.subject_name}",
                'extra': f"{slot.get_day_of_week_display()} {slot.start_time.strftime('%H:%M')}",
                'type': 'Timetable',
                'url': '#',
                'action': f"openTimetable('{slot.timetable.student_class.class_name}')",
            })

    elif scope == 'reports':
        class_id = request.GET.get('class_id')
        if class_id:
            for s in Student.objects.filter(
                enrollments__classroom_id=class_id
            ).filter(full_name_q | Q(admission_number__icontains=q)).distinct()[:8]:
                results.append({
                    'id': s.id,
                    'name': f"{s.first_name} {s.last_name}",
                    'extra': s.admission_number,
                    'type': 'Student',
                    'url': '#',
                    'action': f'highlightStudentRow({s.id})',
                })
        else:
            for s in Student.objects.filter(full_name_q | Q(admission_number__icontains=q))[:8]:
                results.append({
                    'id': s.id,
                    'name': f"{s.first_name} {s.last_name}",
                    'extra': s.admission_number,
                    'type': 'Student',
                    'url': reverse('student_detail', args=[s.id]),
                })

    else:
        # dashboard — global
        for s in Student.objects.filter(full_name_q | Q(admission_number__icontains=q))[:4]:
            results.append({
                'id': s.id,
                'name': f"{s.first_name} {s.last_name}",
                'extra': s.admission_number,
                'type': 'Student',
                'url': reverse('student_detail', args=[s.id]),
            })
        for st in StaffProfile.objects.filter(
            Q(first_name__icontains=q) | Q(last_name__icontains=q) | Q(staff_id__icontains=q)
        )[:4]:
            results.append({
                'id': st.id,
                'name': f"{st.first_name} {st.last_name}",
                'extra': st.staff_id,
                'type': 'Staff',
                'url': reverse('staff_detail', args=[st.id]),
            })
        for p in Parent.objects.filter(
            Q(name__icontains=q) |
            Q(father_of__first_name__icontains=q) |
            Q(father_of__last_name__icontains=q) |
            Q(mother_of__first_name__icontains=q) |
            Q(mother_of__last_name__icontains=q)
        ).distinct()[:3]:
            results.append({
                'id': p.id,
                'name': p.name or 'Unnamed',
                'extra': p.email or '',
                'type': 'Parent',
                'url': reverse('parent_detail', args=[p.id]),
            })
        for c in ClassRoom.objects.filter(class_name__icontains=q)[:3]:
            results.append({
                'id': c.id,
                'name': c.class_name,
                'extra': 'Class',
                'type': 'Class',
                'url': reverse('classes_subjects_hub'),
            })
        for s in Subject.objects.filter(subject_name__icontains=q)[:3]:
            results.append({
                'id': s.id,
                'name': s.subject_name,
                'extra': 'Subject',
                'type': 'Subject',
                'url': reverse('classes_subjects_hub'),
            })

    return JsonResponse({'results': results})


@login_required
def compile_grades_view(request):
    if not _is_staff_or_admin(request.user):
        raise PermissionDenied
    staff = getattr(request.user, 'staff_profile', None)
    if request.user.is_superuser:
        classrooms = ClassRoom.objects.all()
    else:
        assigned_class_ids = StaffClassSubject.objects.filter(staff=staff).values_list('classroom_id', flat=True).distinct()
        classrooms = ClassRoom.objects.filter(id__in=assigned_class_ids)

    selected_class_id = request.GET.get('class_id')
    selected_subject_id = request.GET.get('subject_id')
    assessment_type = request.GET.get('assessment_type', 'class_score')
    selected_subject = None
    classroom = None
    students = []
    available_subjects = Subject.objects.none()

    if selected_class_id:
        classroom = get_object_or_404(ClassRoom, pk=selected_class_id)
        students = Student.objects.filter(enrollments__classroom=classroom).distinct()
        available_subjects = Subject.objects.filter(offered_in_classes__classroom=classroom).distinct()
        if not request.user.is_superuser and staff:
            assigned_subject_ids = StaffClassSubject.objects.filter(staff=staff, classroom=classroom).values_list('subject_id', flat=True).distinct()
            available_subjects = available_subjects.filter(id__in=assigned_subject_ids)

        if selected_subject_id:
            selected_subject = get_object_or_404(Subject, pk=selected_subject_id)

    current_session = AcademicSession.objects.filter(is_current=True).first()
    current_term = Term.objects.filter(is_active=True).first()

    if request.method == 'POST':
        selected_class_id = request.POST.get('class_id')
        selected_subject_id = request.POST.get('subject_id')
        classroom = get_object_or_404(ClassRoom, pk=selected_class_id)
        selected_subject = get_object_or_404(Subject, pk=selected_subject_id)

        if not staff or not StaffClassSubject.objects.filter(staff=staff, classroom=classroom, subject=selected_subject).exists():
            raise PermissionDenied

        students = Student.objects.filter(enrollments__classroom=classroom).distinct()

        for student in students:
            cs = request.POST.get(f'cs_{student.id}')
            es = request.POST.get(f'es_{student.id}')
            if cs or es:
                SubjectAssessment.objects.update_or_create(
                    student=student,
                    subject=selected_subject,
                    academic_session=current_session,
                    academic_term=current_term,
                    defaults={
                        'class_score': cs or 0,
                        'exam_score': es or 0,
                    }
                )
        _notify_form_teacher(request, classroom, selected_subject, False)
        messages.success(request, f'Grades for {selected_subject.subject_name} saved successfully!')
        return redirect(request.path + '?class_id=' + str(selected_class_id) + '&subject_id=' + str(selected_subject_id) + '&assessment_type=' + str(assessment_type))

    grades_matrix = []
    if selected_subject:
        for student in students:
            assessment = SubjectAssessment.objects.filter(
                student=student, subject=selected_subject, academic_session=current_session, academic_term=current_term
            ).first()
            grades_matrix.append({
                'student': student,
                'class_score': assessment.class_score if assessment else '',
                'exam_score': assessment.exam_score if assessment else '',
            })

    context = {
        'classrooms': classrooms,
        'available_subjects': available_subjects,
        'students': students,
        'selected_class': classroom,
        'selected_subject': selected_subject,
        'selected_class_id': selected_class_id,
        'selected_subject_id': selected_subject_id,
        'assessment_type': assessment_type,
        'grades_matrix': grades_matrix,
    }
    return render(request, 'sis/compile_grades.html', context)


@login_required
def midterm_summary_view(request):
    if not _is_staff_or_admin(request.user):
        raise PermissionDenied

    staff = getattr(request.user, 'staff_profile', None)
    classrooms = ClassRoom.objects.none()
    if request.user.is_superuser:
        classrooms = ClassRoom.objects.all()
    elif staff:
        assigned_ids = StaffClassSubject.objects.filter(staff=staff).values_list('classroom_id', flat=True).distinct()
        classrooms = ClassRoom.objects.filter(id__in=assigned_ids)

    selected_class_id = request.GET.get('class_id')
    current_subject_id = request.GET.get('subject_id')

    if selected_class_id and not selected_class_id.isdigit():
        selected_class_id = None
    if current_subject_id and not current_subject_id.isdigit():
        current_subject_id = None

    staff_profile = getattr(request.user, 'staff_profile', None)

    classroom = None
    students = Student.objects.none()
    assigned_subjects = Subject.objects.none()
    report_data = []
    has_records = False

    if selected_class_id:
        classroom = get_object_or_404(ClassRoom, pk=selected_class_id)

        if request.user.is_superuser:
            assigned_subjects = Subject.objects.filter(
                assigned_teachers__classroom_id=selected_class_id
            ).distinct()
        elif staff_profile:
            assigned_subjects = Subject.objects.filter(
                assigned_teachers__staff=staff_profile,
                assigned_teachers__classroom_id=selected_class_id
            ).distinct()

        students = Student.objects.filter(enrollments__classroom=classroom).distinct()

        current_session = AcademicSession.objects.filter(is_current=True).first()
        current_term = Term.objects.filter(is_active=True).first()

        for student in students:
            records = MidTermRecord.objects.filter(student=student, classroom=classroom)
            if current_session:
                records = records.filter(academic_session=current_session)
            if current_term:
                records = records.filter(term=current_term)
            if current_subject_id:
                records = records.filter(subject_id=current_subject_id)

            first_record = records.first()
            midterm_score = (
                f"{first_record.midterm_score:.1f}"
                if first_record and first_record.midterm_score is not None
                else None
            )
            total = sum(
                r.midterm_score for r in records if r.midterm_score is not None
            )
            row = {
                'student': student,
                'midterm_score': midterm_score,
                'total': total,
            }
            report_data.append(row)

        has_records = bool(
            report_data and any(
                row['midterm_score'] is not None or row['total'] > 0
                for row in report_data
            )
        )
        report_data.sort(key=lambda x: x['total'], reverse=True)
        for idx, row in enumerate(report_data):
            row['rank'] = idx + 1

    try:
        user_form_class = request.user.staff_profile.form_class
    except AttributeError:
        user_form_class = None

    context = {
        'classrooms': classrooms,
        'classroom': classroom,
        'selected_class_id': selected_class_id,
        'assigned_subjects': assigned_subjects,
        'current_subject_id': (
            int(current_subject_id)
            if current_subject_id and current_subject_id.isdigit()
            else None
        ),
        'students': students,
        'report_data': report_data,
        'has_records': has_records,
        'user_form_class': user_form_class,
    }
    return render(request, 'sis/midterm_summary.html', context)


@login_required
def compile_midterm_grades_view(request):
    if not _is_staff_or_admin(request.user):
        raise PermissionDenied

    staff = getattr(request.user, 'staff_profile', None)

    if request.user.is_superuser:
        classrooms = ClassRoom.objects.all()
    else:
        assigned_class_ids = StaffClassSubject.objects.filter(staff=staff).values_list('classroom_id', flat=True).distinct()
        classrooms = ClassRoom.objects.filter(id__in=assigned_class_ids)

    selected_class_id = request.GET.get('class_id')
    selected_subject_id = request.GET.get('subject_id')
    selected_subject = None
    auto_selected_subject = False
    classroom = None
    students = []
    available_subjects = Subject.objects.none()

    current_session = AcademicSession.objects.filter(is_current=True).first()
    current_term = Term.objects.filter(is_active=True).first()

    if selected_class_id:
        classroom = get_object_or_404(ClassRoom, pk=selected_class_id)
        students = Student.objects.filter(enrollments__classroom=classroom).distinct()
        available_subjects = Subject.objects.filter(offered_in_classes__classroom=classroom).distinct()
        if not request.user.is_superuser and staff:
            assigned_subject_ids = StaffClassSubject.objects.filter(staff=staff, classroom=classroom).values_list('subject_id', flat=True).distinct()
            available_subjects = available_subjects.filter(id__in=assigned_subject_ids)

        if selected_subject_id:
            selected_subject = get_object_or_404(Subject, pk=selected_subject_id)
        else:
            last_record = MidTermRecord.objects.filter(
                classroom=classroom,
                academic_session=current_session,
                term=current_term,
            ).order_by('-date_recorded').first()
            if last_record and available_subjects.filter(pk=last_record.subject_id).exists():
                selected_subject = last_record.subject
                selected_subject_id = str(selected_subject.id)
                auto_selected_subject = True

    if request.method == 'POST':
        selected_class_id = request.POST.get('class_id')
        selected_subject_id = request.POST.get('subject_id')
        classroom = get_object_or_404(ClassRoom, pk=selected_class_id)
        selected_subject = get_object_or_404(Subject, pk=selected_subject_id)

        if not staff or not StaffClassSubject.objects.filter(staff=staff, classroom=classroom, subject=selected_subject).exists():
            raise PermissionDenied

        students = Student.objects.filter(enrollments__classroom=classroom).distinct()

        for student in students:
            val = request.POST.get(f'midterm_score_{student.id}')
            if val:
                MidTermRecord.objects.update_or_create(
                    student=student,
                    academic_session=current_session,
                    term=current_term,
                    subject=selected_subject,
                    defaults={
                        'classroom': classroom,
                        'midterm_score': val,
                        'recorded_by': staff,
                    }
                )

        messages.success(request, f'Mid-term grades for {selected_subject.subject_name} saved successfully!')
        return redirect(request.path + '?class_id=' + str(selected_class_id) + '&subject_id=' + str(selected_subject.id))

    grade_map = {}
    if selected_subject:
        records = MidTermRecord.objects.filter(
            student__in=students,
            subject=selected_subject,
            academic_session=current_session,
            term=current_term,
        )
        grade_map = {rec.student_id: rec.midterm_score for rec in records}

    ranks = {}
    if selected_subject:
        scored = sorted(
            [(student.id, (grade_map.get(student.id) or 0)) for student in students],
            key=lambda x: x[1],
            reverse=True,
        )
        dense_rank = 0
        prev_score = None
        for student_id, score in scored:
            if score != prev_score:
                dense_rank += 1
                prev_score = score
            ranks[student_id] = dense_rank

    context = {
        'classrooms': classrooms,
        'available_subjects': available_subjects,
        'students': students,
        'selected_class': classroom,
        'selected_subject': selected_subject,
        'selected_class_id': selected_class_id,
        'selected_subject_id': selected_subject_id,
        'subject_selected': bool(selected_subject),
        'auto_selected_subject': auto_selected_subject,
        'grade_map': grade_map,
        'ranks': ranks,
    }
    return render(request, 'sis/compile_midterm_grades.html', context)


@login_required
def midterm_generate_report_hub_view(request, class_id):
    if not _is_staff_or_admin(request.user):
        raise PermissionDenied
    classroom = ClassRoom.objects.get(pk=class_id)
    staff = getattr(request.user, 'staff_profile', None)
    is_form_teacher = staff and staff.form_class == classroom
    has_full_access = request.user.is_superuser or is_form_teacher
    if not has_full_access:
        raise PermissionDenied

    current_session = AcademicSession.objects.filter(is_current=True).first()
    current_term = Term.objects.filter(is_active=True).first()

    subjects_for_class = Subject.objects.filter(offered_in_classes__classroom=classroom).distinct().order_by('subject_name')
    students = Student.objects.filter(enrollments__classroom=classroom).distinct()

    all_records = MidTermRecord.objects.filter(
        academic_session=current_session, academic_term=current_term,
        student__in=students, subject__in=subjects_for_class,
    )
    record_map = {}
    for r in all_records:
        record_map.setdefault(r.student_id, {})[r.subject_id] = r

    GRADE_REMARKS = get_grade_remarks_for_class(classroom)

    def get_remark(total):
        if total is None:
            return "—", "—"
        for floor, grade, label in GRADE_REMARKS:
            if total >= floor:
                return grade, label
        return GRADE_REMARKS[-1][1], GRADE_REMARKS[-1][2]

    report_data = []
    for student in students:
        subject_scores = {}
        for subj in subjects_for_class:
            rec = record_map.get(student.id, {}).get(subj.id)
            if rec and rec.midterm_score is not None:
                score = float(rec.midterm_score)
                grade, remark = get_remark(score)
                subject_scores[subj.id] = {
                    'score': score,
                    'grade': grade,
                    'remark': remark,
                }
            else:
                subject_scores[subj.id] = {
                    'score': None,
                    'grade': '—',
                    'remark': '—',
                }

        total = sum(
            s['score'] for s in subject_scores.values() if s['score'] is not None
        )
        grade, remark = get_remark(total)

        report_data.append({
            'student': student,
            'subject_scores': subject_scores,
            'total': total,
            'overall_grade': grade,
            'overall_remark': remark,
        })

    report_data = sorted(report_data, key=lambda x: x['total'], reverse=True)
    for index, row in enumerate(report_data):
        row['rank'] = index + 1
        row['rank_ordinal'] = get_ordinal(row['rank'])

    subject_positions = {}
    for subj in subjects_for_class:
        scored = [(r['student'].id, r['subject_scores'][subj.id]['score'])
                  for r in report_data
                  if r['subject_scores'][subj.id]['score'] is not None]
        scored.sort(key=lambda x: x[1], reverse=True)
        for idx, (sid, _) in enumerate(scored):
            subject_positions.setdefault(subj.id, {})[sid] = idx + 1

    if request.user.is_superuser:
        classrooms = ClassRoom.objects.all()
    elif staff:
        assigned_ids = StaffClassSubject.objects.filter(staff=staff).values_list('classroom_id', flat=True).distinct()
        classrooms = ClassRoom.objects.filter(id__in=assigned_ids)
    else:
        classrooms = ClassRoom.objects.none()

    return render(request, 'sis/midterm_generate_report_hub.html', {
        'classroom': classroom,
        'classrooms': classrooms,
        'report_data': report_data,
        'subjects_for_class': subjects_for_class,
        'subject_positions': subject_positions,
        'current_session': current_session,
        'current_term': current_term,
        'term_number': int(current_term.term_name.split()[-1]) if current_term else 1,
        'year_label': current_session.academic_year if current_session else '',
        'student_count': len(report_data),
    })


@login_required
def midterm_print_report_cards_view(request, class_id):
    if not _is_staff_or_admin(request.user):
        raise PermissionDenied
    classroom = ClassRoom.objects.get(pk=class_id)
    staff = getattr(request.user, 'staff_profile', None)
    is_form_teacher = staff and staff.form_class == classroom
    has_full_access = request.user.is_superuser or is_form_teacher
    if not has_full_access:
        raise PermissionDenied

    student_ids = request.POST.getlist('student_ids[]')
    if not student_ids:
        return render(request, 'sis/midterm_report_card_print.html', {
            'classroom': classroom,
            'selected_students': [],
            'subjects_for_class': [],
            'current_session': None,
            'current_term': None,
            'term_number': 1,
            'year_label': '',
        })

    current_session = AcademicSession.objects.filter(is_current=True).first()
    current_term = Term.objects.filter(is_active=True).first() if current_session else None

    subjects_for_class = Subject.objects.filter(offered_in_classes__classroom=classroom).distinct().order_by('subject_name')
    
    # Get all students for ranking, not just selected ones
    all_students = Student.objects.filter(enrollments__classroom=classroom).distinct()
    selected_student_ids = set(int(sid) for sid in student_ids)

    all_records = MidTermRecord.objects.filter(
        academic_session=current_session, academic_term=current_term,
        student__in=all_students, subject__in=subjects_for_class,
    )
    record_map = {}
    for r in all_records:
        record_map.setdefault(r.student_id, {})[r.subject_id] = r

    GRADE_REMARKS = get_grade_remarks_for_class(classroom)

    def get_remark(total):
        if total is None:
            return "—", "—"
        for floor, grade, label in GRADE_REMARKS:
            if total >= floor:
                return grade, label
        return GRADE_REMARKS[-1][1], GRADE_REMARKS[-1][2]

    report_data = []
    # Build report for all students, then filter to selected
    for student in all_students:
        subject_scores = {}
        for subj in subjects_for_class:
            rec = record_map.get(student.id, {}).get(subj.id)
            if rec and rec.midterm_score is not None:
                score = float(rec.midterm_score)
                grade, remark = get_remark(score)
                subject_scores[subj.id] = {
                    'subject': subj,
                    'score': score,
                    'grade': grade,
                    'remark': remark,
                }
            else:
                subject_scores[subj.id] = {
                    'subject': subj,
                    'score': None,
                    'grade': '—',
                    'remark': '—',
                }

        total = sum(
            s['score'] for s in subject_scores.values() if s['score'] is not None
        )
        grade, remark = get_remark(total)

        report_data.append({
            'student': student,
            'subject_scores': subject_scores,
            'total': total,
            'overall_grade': grade,
            'overall_remark': remark,
        })

    report_data = sorted(report_data, key=lambda x: x['total'], reverse=True)
    # Rank all students
    for index, row in enumerate(report_data):
        row['rank'] = index + 1
        row['rank_ordinal'] = get_ordinal(row['rank'])

    subject_positions = {}
    for subj in subjects_for_class:
        scored = [(r['student'].id, r['subject_scores'][subj.id]['score'])
                  for r in report_data
                  if r['subject_scores'][subj.id]['score'] is not None]
        scored.sort(key=lambda x: x[1], reverse=True)
        for idx, (sid, _) in enumerate(scored):
            subject_positions.setdefault(subj.id, {})[sid] = idx + 1

    for row in report_data:
        for subj in subjects_for_class:
            row['subject_scores'][subj.id]['subject_position'] = subject_positions.get(subj.id, {}).get(row['student'].id)

    # Filter to only selected students for printing
    selected_students = [r for r in report_data if r['student'].id in selected_student_ids]

    class_size = all_students.count()

    form_master_name = ""
    if classroom.form_master:
        form_master_name = f"{classroom.form_master.first_name} {classroom.form_master.last_name}"

    return render(request, 'sis/midterm_report_card_print.html', {
        'classroom': classroom,
        'selected_students': selected_students,
        'subjects_for_class': subjects_for_class,
        'current_session': current_session,
        'current_term': current_term,
        'term_number': int(current_term.term_name.split()[-1]) if current_term else 1,
        'year_label': current_session.academic_year if current_session else '',
        'class_size': class_size,
        'form_master_name': form_master_name,
    })


@login_required
def api_class_subjects(request):
    class_id = request.GET.get('class_id')
    if not class_id:
        return JsonResponse({'subjects': []})
    mappings = ClassSubject.objects.filter(classroom_id=class_id).select_related('subject')
    subjects_list = []
    for m in mappings:
        existing = StaffClassSubject.objects.filter(
            classroom_id=class_id,
            subject=m.subject
        ).select_related('staff__user').first()
        is_assigned = existing is not None
        assigned_teacher_name = None
        if is_assigned:
            assigned_teacher_name = existing.staff.user.get_full_name() if existing.staff.user else f"{existing.staff.first_name} {existing.staff.last_name}"
        subjects_list.append({
            'id': m.subject.id,
            'name': m.subject.subject_name,
            'is_assigned': is_assigned,
            'assigned_teacher_name': assigned_teacher_name,
        })
    return JsonResponse({'subjects': subjects_list})


@login_required
def api_class_details(request, class_id):
    classroom = get_object_or_404(ClassRoom, pk=class_id)
    student_count = Student.objects.filter(enrollments__classroom=classroom).distinct().count()
    form_teacher_name = ''
    if classroom.form_teacher:
        ft = classroom.form_teacher
        form_teacher_name = f"{ft.first_name} {ft.last_name}".strip()

    subjects_data = []
    cs_qs = ClassSubject.objects.filter(classroom=classroom).select_related('subject')
    for cs in cs_qs:
        assignment = StaffClassSubject.objects.filter(
            classroom=classroom, subject=cs.subject
        ).select_related('staff__user').first()
        teacher_name = ''
        if assignment and assignment.staff:
            teacher_name = f"{assignment.staff.first_name} {assignment.staff.last_name}".strip()
        subjects_data.append({
            'name': cs.subject.subject_name,
            'teacher': teacher_name or 'Unassigned',
        })

    return JsonResponse({
        'id': classroom.id,
        'name': classroom.class_name,
        'student_count': student_count,
        'form_teacher': form_teacher_name or 'Unassigned',
        'subjects': subjects_data,
    })


@login_required
def api_subject_details(request, subject_id):
    subject = get_object_or_404(Subject, pk=subject_id)

    class_subjects = ClassSubject.objects.filter(subject=subject).select_related('classroom')

    scs_qs = StaffClassSubject.objects.filter(subject=subject).select_related('staff')

    teacher_map = {}
    all_teachers = set()
    for scs in scs_qs:
        c_id = scs.classroom_id
        t_name = ''
        if scs.staff:
            t_name = f"{scs.staff.first_name} {scs.staff.last_name}".strip()
        if t_name:
            all_teachers.add(t_name)
            teacher_map.setdefault(c_id, [])
            if t_name not in teacher_map[c_id]:
                teacher_map[c_id].append(t_name)

    class_allocations = []
    for cs in class_subjects:
        c_id = cs.classroom.id
        teachers = teacher_map.get(c_id, [])
        class_allocations.append({
            'class_name': cs.classroom.class_name,
            'teacher': ', '.join(teachers) if teachers else 'Unassigned',
        })

    return JsonResponse({
        'id': subject.id,
        'name': subject.subject_name,
        'category': 'General Education',
        'classes_count': len(class_allocations),
        'assigned_teachers': sorted(all_teachers),
        'class_allocations': class_allocations,
    })


@login_required
def verify_class_rankings_view(request, class_id):
    if not _is_staff_or_admin(request.user):
        raise PermissionDenied

    classroom = get_object_or_404(ClassRoom, pk=class_id)

    staff = getattr(request.user, 'staff_profile', None)
    is_form_teacher = staff and staff.form_class == classroom

    if not is_form_teacher and not _is_admin(request.user):
        messages.error(request, 'Only the form teacher of this class can verify rankings.')
        return redirect('class_report', class_id=class_id)

    students = Student.objects.filter(enrollments__classroom=classroom).distinct()

    current_session = AcademicSession.objects.filter(is_current=True).first()
    current_term = Term.objects.filter(is_active=True).first()

    report_data = []
    for student in students:
        assessments = SubjectAssessment.objects.filter(student=student, academic_session=current_session, academic_term=current_term)
        grand_total = sum(ast.total_score for ast in assessments)
        report_data.append({
            'student': student,
            'assessments': assessments,
            'grand_total': grand_total,
        })

    report_data = sorted(report_data, key=lambda x: x['grand_total'], reverse=True)
    for index, row in enumerate(report_data):
        row['rank'] = index + 1

    verification = GradeVerification.objects.filter(
        classroom=classroom, academic_session=current_session, academic_term=current_term
    ).first()

    if request.method == 'POST':
        if not verification:
            GradeVerification.objects.create(
                classroom=classroom,
                verified_by=staff,
                academic_session=current_session,
                academic_term=current_term,
            )
            messages.success(request, f'Rankings for {classroom.class_name} verified successfully.')
        else:
            messages.info(request, 'Rankings were already verified for this term.')
        return redirect('class_report', class_id=class_id)

    has_graded_records = any(row['assessments'].exists() for row in report_data)

    return render(request, 'sis/class_report.html', {
        'classroom': classroom,
        'report_data': report_data,
        'is_form_teacher': is_form_teacher,
        'verification': verification,
        'has_graded_records': has_graded_records,
    })


@login_required
def view_account(request):
    staff_profile = getattr(request.user, 'staff_profile', None)
    subjects_with_classes = []
    my_students = []

    if staff_profile:
        assignments = StaffClassSubject.objects.filter(staff=staff_profile).select_related('classroom', 'subject')
        subjects_with_classes = assignments

        if staff_profile.form_class:
            my_students = Student.objects.filter(classroom=staff_profile.form_class).order_by('first_name')
        else:
            class_ids = assignments.values_list('classroom_id', flat=True).distinct()
            my_students = Student.objects.filter(classroom_id__in=class_ids).order_by('first_name') if class_ids else []

    return render(request, 'sis/account.html', {
        'staff_profile': staff_profile,
        'subjects_with_classes': subjects_with_classes,
        'my_students': my_students,
    })


@login_required
def parents_list(request):
    user = request.user

    children_prefetch = Prefetch(
        'father_of',
        queryset=Student.objects.select_related('classroom').filter(is_alumni=False),
        to_attr='father_children'
    )
    mother_prefetch = Prefetch(
        'mother_of',
        queryset=Student.objects.select_related('classroom').filter(is_alumni=False),
        to_attr='mother_children'
    )

    if user.is_superuser:
        parents = Parent.objects.prefetch_related(children_prefetch, mother_prefetch).order_by('name')
    elif hasattr(user, 'staff_profile'):
        staff = user.staff_profile
        classroom_ids = StaffClassSubject.objects.filter(
            staff=staff
        ).values_list('classroom_id', flat=True).distinct()
        parents = Parent.objects.filter(
            Q(father_of__classroom_id__in=classroom_ids) |
            Q(mother_of__classroom_id__in=classroom_ids)
        ).distinct().prefetch_related(children_prefetch, mother_prefetch).order_by('name')
    else:
        parents = Parent.objects.none()

    return render(request, 'sis/parents_list.html', {'parents': parents})


@login_required
def export_session_excel(request):
    if not _is_admin(request.user):
        raise PermissionDenied

    from datetime import datetime
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    current_session = AcademicSession.objects.filter(is_current=True).first()
    current_term = Term.objects.filter(is_active=True).first()

    header_font = Font(bold=True, size=14)
    sub_header_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="E8E8E8", end_color="E8E8E8", fill_type="solid")

    def set_headers(ws, headers):
        for col_idx, name in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=name)
            cell.font = sub_header_font
            cell.border = border
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

    def write_row(ws, values, row):
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.border = border
        return row + 1

    def autofit(ws):
        for col in ws.columns:
            max_length = 0
            col_letter = None
            for cell in col:
                if not hasattr(cell, 'column_letter'):
                    continue
                if col_letter is None:
                    col_letter = cell.column_letter
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            if col_letter:
                ws.column_dimensions[col_letter].width = min(max_length + 4, 45)

    def sheet_title(ws, title, ncols):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        ws.cell(row=1, column=1, value=title).font = header_font

    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"
    sheet_title(ws, "SESSION DATABASE — OVERVIEW", 2)
    set_headers(ws, ["Field", "Value"])

    all_sessions = AcademicSession.objects.all()
    all_terms = Term.objects.all()
    overview = [
        ("Database Exported", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ("Academic Session (Year)", current_session.academic_year if current_session else "None"),
        ("Active Term", current_term.term_name if current_term else "None"),
        ("All Sessions", all_sessions.count()),
        ("All Terms", all_terms.count()),
        ("Total Classes", ClassRoom.objects.count()),
        ("Total Subjects", Subject.objects.count()),
        ("Total Departments", Department.objects.count()),
        ("Total Designations", Designation.objects.count()),
        ("Total Parents", Parent.objects.count()),
        ("Total Students (all)", Student.objects.count()),
        ("Non-Alumni Students", Student.objects.filter(is_alumni=False).count()),
        ("Total Enrollments", Enrollment.objects.count()),
        ("Subject Assessments (all)", SubjectAssessment.objects.count()),
        ("Mid-Term Records (all)", MidTermRecord.objects.count()),
        ("Grade Verifications (all)", GradeVerification.objects.count()),
        ("Total Staff", StaffProfile.objects.count()),
        ("Staff Class-Subject Assignments", StaffClassSubject.objects.count()),
        ("Promotion Criteria", PromotionCriteria.objects.count()),
        ("Notifications", Notification.objects.count()),
        ("Timetables", Timetable.objects.count()),
        ("Timetable Slots", TimetableSlot.objects.count()),
    ]
    row = 4
    for field, value in overview:
        row = write_row(ws, [field, value], row)
    autofit(ws)

    def add_table(sheet_name, title, headers, rows):
        ws = wb.create_sheet(sheet_name)
        sheet_title(ws, title, len(headers))
        set_headers(ws, headers)
        r = 4
        for values in rows:
            r = write_row(ws, values, r)
        autofit(ws)
        return ws

    add_table("Sessions", "ACADEMIC SESSIONS", ["ID", "Academic Year", "Is Current"], [
        [s.pk, s.academic_year, s.is_current] for s in all_sessions.order_by('academic_year')
    ])

    add_table("Terms", "TERMS", ["ID", "Session", "Term Name", "Is Active"], [
        [t.pk, t.session.academic_year, t.term_name, t.is_active]
        for t in all_terms.select_related('session').order_by('session__academic_year', 'id')
    ])

    add_table("Classes", "CLASSES", ["ID", "Class Name", "Order", "Next Class", "Form Master"], [
        [c.pk, c.class_name, c.order, c.next_class.class_name if c.next_class else '',
         c.form_master.staff_id if c.form_master else '']
        for c in ClassRoom.objects.select_related('next_class', 'form_master')
    ])

    add_table("Subjects", "SUBJECTS", ["ID", "Subject Name"], [
        [s.pk, s.subject_name] for s in Subject.objects.order_by('subject_name')
    ])

    add_table("Departments", "DEPARTMENTS", ["ID", "Name"], [
        [d.pk, d.name] for d in Department.objects.all()
    ])

    add_table("Designations", "DESIGNATIONS", ["ID", "Name"], [
        [d.pk, d.name] for d in Designation.objects.all()
    ])

    add_table("Class Subjects", "CLASS SUBJECTS (SUBJECTS OFFERED PER CLASS)", [
        "ID", "Class", "Subject",
    ], [
        [cs.pk, cs.classroom.class_name, cs.subject.subject_name]
        for cs in ClassSubject.objects.select_related('classroom', 'subject').order_by('classroom__order', 'subject__subject_name')
    ])

    add_table("Parents", "PARENTS", [
        "ID", "Name", "Occupation", "Residential Address", "Email", "Telephone",
        "No. of Children",
    ], [
        [p.pk, p.name or '', p.occupation or '', p.residential_address or '',
         p.email or '', p.telephone_number or '',
         p.father_of.count() + p.mother_of.count()]
        for p in Parent.objects.all().order_by('name')
    ])

    add_table("Students", "STUDENTS", [
        "ID", "Admission No.", "First Name", "Other Names", "Last Name", "Gender",
        "Date of Birth", "Date of Admission", "Status", "Living With",
        "Previous School", "Class", "Father Name", "Father Phone", "Father Email",
        "Mother Name", "Mother Phone", "Mother Email",
        "Pending Next Class", "Promotion Status", "Is Alumni",
    ], [
        [
            s.id, s.admission_number, s.first_name, s.other_names or '', s.last_name,
            s.gender,
            s.dob.strftime('%Y-%m-%d') if s.dob else '',
            s.date_of_admission.strftime('%Y-%m-%d') if s.date_of_admission else '',
            s.status, s.living_with, s.previous_school_attended or '',
            s.classroom.class_name if s.classroom else '',
            s.father.name if s.father else '', s.father.telephone_number if s.father else '',
            s.father.email if s.father else '', s.mother.name if s.mother else '',
            s.mother.telephone_number if s.mother else '', s.mother.email if s.mother else '',
            s.pending_next_class.class_name if s.pending_next_class else '',
            s.promotion_status, s.is_alumni,
        ]
        for s in Student.objects.select_related(
            'classroom', 'father', 'mother', 'pending_next_class'
        ).order_by('last_name', 'first_name')
    ])

    add_table("Enrollments", "ENROLLMENTS", [
        "ID", "Student", "Admission No.", "Class", "Academic Session", "Term",
        "Date Enrolled",
    ], [
        [e.pk, e.student.get_full_name, e.student.admission_number,
         e.classroom.class_name, e.academic_session.academic_year,
         e.academic_term.term_name,
         e.date_enrolled.strftime('%Y-%m-%d %H:%M') if e.date_enrolled else '']
        for e in Enrollment.objects.select_related(
            'student', 'classroom', 'academic_session', 'academic_term'
        ).order_by('academic_session__academic_year', 'student__last_name', 'student__first_name')
    ])

    add_table("Subject Assessments", "SUBJECT ASSESSMENTS", [
        "ID", "Student", "Admission No.", "Subject", "Academic Session", "Term",
        "Class Score (30%)", "Exam Score (70%)", "Total", "Grade", "Remark", "Is Active Session/Term",
    ], [
        [a.pk, a.student.get_full_name, a.student.admission_number,
         a.subject.subject_name, a.academic_session.academic_year,
         a.academic_term.term_name, float(a.class_score), float(a.exam_score),
         a.total_score, a.grade_and_remark[0], a.grade_and_remark[1],
         bool(current_session and current_term and
              a.academic_session_id == current_session.pk and a.academic_term_id == current_term.pk)]
        for a in SubjectAssessment.objects.select_related(
            'student', 'subject', 'academic_session', 'academic_term'
        ).order_by('academic_session__academic_year', 'student__last_name', 'student__first_name', 'subject__subject_name')
    ])

    add_table("Mid-Term Records", "MID-TERM RECORDS", [
        "ID", "Student", "Admission No.", "Class", "Subject", "Academic Session",
        "Term", "Mid-Term Score", "Recorded By", "Date Recorded",
    ], [
        [m.pk, m.student.get_full_name, m.student.admission_number,
         m.classroom.class_name if m.classroom else '',
         m.subject.subject_name, m.academic_session.academic_year, m.term.term_name,
         float(m.midterm_score) if m.midterm_score is not None else '',
         m.recorded_by.staff_id if m.recorded_by else '',
         m.date_recorded.strftime('%Y-%m-%d %H:%M') if m.date_recorded else '']
        for m in MidTermRecord.objects.select_related(
            'student', 'classroom', 'subject', 'academic_session', 'term', 'recorded_by'
        ).order_by('academic_session__academic_year', 'student__last_name', 'student__first_name')
    ])

    add_table("Grade Verifications", "GRADE VERIFICATIONS", [
        "ID", "Class", "Academic Session", "Term", "Verified By", "Verified At",
    ], [
        [g.pk, g.classroom.class_name, g.academic_session.academic_year,
         g.academic_term.term_name, g.verified_by.staff_id if g.verified_by else '',
         g.verified_at.strftime('%Y-%m-%d %H:%M') if g.verified_at else '']
        for g in GradeVerification.objects.select_related('classroom', 'verified_by', 'academic_session', 'academic_term')
    ])

    add_table("Staff", "STAFF", [
        "ID", "Staff ID", "Username", "Title", "First Name", "Other Names", "Last Name",
        "Gender", "Date of Birth", "Designation", "Department", "SSNIT ID", "Phone",
        "Email", "Employment Type", "Date of Appointment", "Year of Last Promotion",
        "Qualification", "Certificate", "Institution Completed", "Year Completed",
        "Form Class (Form Teacher)", "Residential Address", "Theme",
        "Active Account", "Superuser / Staff",
    ], [
        [s.pk, s.staff_id, s.user.username if s.user else '',
         s.title, s.first_name, s.other_names or '', s.last_name, s.gender,
         s.dob.strftime('%Y-%m-%d') if s.dob else '',
         s.designation.name if s.designation else '',
         s.department.name if s.department else '', s.ssnit_id or '', s.phone_number or '',
         s.email, s.employment_type,
         s.date_of_appointment.strftime('%Y-%m-%d') if s.date_of_appointment else '',
         s.year_of_last_promotion if s.year_of_last_promotion is not None else '',
         s.qualification, s.certificate, s.name_of_institution_completed, s.year_completed,
         s.form_class.class_name if s.form_class else '', s.address or '', s.theme,
         s.user.is_active if s.user else False,
         s.user.is_superuser or s.user.is_staff if s.user else False]
        for s in StaffProfile.objects.select_related(
            'user', 'designation', 'department', 'form_class'
        ).order_by('last_name', 'first_name')
    ])

    add_table("Staff Assignments", "STAFF CLASS-SUBJECT ASSIGNMENTS", [
        "ID", "Staff ID", "Staff Name", "Class", "Subject",
    ], [
        [scs.pk, scs.staff.staff_id, f"{scs.staff.first_name} {scs.staff.last_name}".strip(),
         scs.classroom.class_name, scs.subject.subject_name]
        for scs in StaffClassSubject.objects.select_related(
            'staff', 'classroom', 'subject'
        ).order_by('staff__staff_id', 'classroom__order')
    ])

    add_table("Promotion Criteria", "PROMOTION CRITERIA", [
        "ID", "Class", "Minimum Grand Total",
    ], [
        [pc.pk, pc.classroom.class_name, float(pc.min_grand_total)]
        for pc in PromotionCriteria.objects.select_related('classroom')
    ])

    add_table("Notifications", "NOTIFICATIONS", [
        "ID", "Recipient", "Type", "Title", "Message", "Is Read", "Created At",
    ], [
        [n.pk, n.recipient.username, n.notification_type, n.title, n.message, n.is_read,
         n.created_at.strftime('%Y-%m-%d %H:%M') if n.created_at else '']
        for n in Notification.objects.select_related('recipient')
    ])

    add_table("Timetables", "TIMETABLES", [
        "ID", "Class", "Title", "Term", "Is Active", "Created At", "Updated At",
    ], [
        [t.pk, t.student_class.class_name, t.title, str(t.academic_term),
         t.is_active,
         t.created_at.strftime('%Y-%m-%d %H:%M') if t.created_at else '',
         t.updated_at.strftime('%Y-%m-%d %H:%M') if t.updated_at else '']
        for t in Timetable.objects.select_related('student_class', 'academic_term')
    ])

    add_table("Timetable Slots", "TIMETABLE SLOTS", [
        "ID", "Timetable", "Class", "Subject", "Teacher", "Day", "Start Time",
        "End Time", "Room / Note",
    ], [
        [ts.pk, f"{ts.timetable.student_class.class_name} - {ts.timetable.title}",
         ts.timetable.student_class.class_name, ts.subject.subject_name,
         ts.teacher.staff_id if ts.teacher else '',
         ts.get_day_of_week_display(),
         ts.start_time.strftime('%H:%M:%S') if ts.start_time else '',
         ts.end_time.strftime('%H:%M:%S') if ts.end_time else '',
         ts.room_or_note or '']
        for ts in TimetableSlot.objects.select_related('timetable', 'subject', 'teacher')
    ])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    year_label = current_session.academic_year.replace('/', '-') if current_session else 'unknown'
    term_label = current_term.term_name.replace(' ', '') if current_term else 'unknown'
    filename = f"session_database_{year_label}_{term_label}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def export_session_json(request):
    if not _is_admin(request.user):
        raise PermissionDenied

    import json
    from datetime import datetime

    current_session = AcademicSession.objects.filter(is_current=True).first()
    current_term = Term.objects.filter(is_active=True).first()

    payload = {
        "exported_at": datetime.now().isoformat(),
        "session": {
            "academic_year": current_session.academic_year if current_session else None,
            "is_current": current_session.is_current if current_session else False,
        },
        "active_term": {
            "term_name": current_term.term_name if current_term else None,
            "is_active": current_term.is_active if current_term else False,
        },
        "sessions": [
            {
                "academic_year": s.academic_year,
                "is_current": s.is_current,
            }
            for s in AcademicSession.objects.all().order_by('academic_year')
        ],
        "terms": [
            {
                "session": t.session.academic_year,
                "term_name": t.term_name,
                "is_active": t.is_active,
            }
            for t in Term.objects.select_related('session').order_by('session__academic_year', 'id')
        ],
        "classes": [
            {
                "class_name": cl.class_name,
                "order": cl.order,
                "next_class": cl.next_class.class_name if cl.next_class else None,
                "form_master_id": cl.form_master.staff_id if cl.form_master else None,
            }
            for cl in ClassRoom.objects.select_related('next_class', 'form_master')
        ],
        "subjects": [s.subject_name for s in Subject.objects.all().order_by('subject_name')],
        "departments": [d.name for d in Department.objects.all()],
        "designations": [d.name for d in Designation.objects.all()],
        "class_subjects": [
            {
                "classroom": cs.classroom.class_name,
                "subject": cs.subject.subject_name,
            }
            for cs in ClassSubject.objects.select_related('classroom', 'subject')
        ],
        "parents": [
            {
                "id": p.pk,
                "name": p.name or '',
                "occupation": p.occupation or '',
                "residential_address": p.residential_address or '',
                "email": p.email or '',
                "telephone_number": p.telephone_number or '',
            }
            for p in Parent.objects.all().order_by('name')
        ],
        "students": [
            {
                "id": s.pk,
                "admission_number": s.admission_number,
                "first_name": s.first_name,
                "other_names": s.other_names or '',
                "last_name": s.last_name,
                "gender": s.gender,
                "dob": s.dob.isoformat() if s.dob else None,
                "date_of_admission": s.date_of_admission.isoformat() if s.date_of_admission else None,
                "status": s.status,
                "living_with": s.living_with,
                "previous_school_attended": s.previous_school_attended or '',
                "classroom": s.classroom.class_name if s.classroom else None,
                "father_id": s.father_id,
                "mother_id": s.mother_id,
                "pending_next_class": s.pending_next_class.class_name if s.pending_next_class else None,
                "promotion_status": s.promotion_status,
                "is_alumni": s.is_alumni,
            }
            for s in Student.objects.select_related(
                'classroom', 'father', 'mother', 'pending_next_class'
            ).order_by('last_name', 'first_name')
        ],
        "enrollments": [
            {
                "student": e.student.admission_number,
                "classroom": e.classroom.class_name,
                "academic_session": e.academic_session.academic_year,
                "academic_term": e.academic_term.term_name,
                "date_enrolled": e.date_enrolled.isoformat() if e.date_enrolled else None,
            }
            for e in Enrollment.objects.select_related(
                'student', 'classroom', 'academic_session', 'academic_term'
            ).order_by('academic_session__academic_year', 'student__last_name', 'student__first_name')
        ],
        "subject_assessments": [
            {
                "student": a.student.admission_number,
                "subject": a.subject.subject_name,
                "academic_session": a.academic_session.academic_year,
                "academic_term": a.academic_term.term_name,
                "class_score": float(a.class_score),
                "exam_score": float(a.exam_score),
                "total": a.total_score,
                "grade": a.grade_and_remark[0],
                "remark": a.grade_and_remark[1],
                "is_active_session_term": bool(
                    current_session and current_term and
                    a.academic_session_id == current_session.pk and a.academic_term_id == current_term.pk
                ),
            }
            for a in SubjectAssessment.objects.select_related(
                'student', 'subject', 'academic_session', 'academic_term'
            ).order_by('academic_session__academic_year', 'student__last_name', 'student__first_name', 'subject__subject_name')
        ],
        "midterm_records": [
            {
                "student": m.student.admission_number,
                "subject": m.subject.subject_name,
                "academic_session": m.academic_session.academic_year,
                "term": m.term.term_name,
                "classroom": m.classroom.class_name if m.classroom else None,
                "midterm_score": float(m.midterm_score) if m.midterm_score is not None else None,
                "recorded_by": m.recorded_by.staff_id if m.recorded_by else None,
            }
            for m in MidTermRecord.objects.select_related(
                'student', 'subject', 'academic_session', 'term', 'classroom', 'recorded_by'
            ).order_by('academic_session__academic_year', 'student__last_name', 'student__first_name')
        ],
        "grade_verifications": [
            {
                "classroom": gv.classroom.class_name,
                "academic_session": gv.academic_session.academic_year,
                "academic_term": gv.academic_term.term_name,
                "verified_by": gv.verified_by.staff_id if gv.verified_by else None,
            }
            for gv in GradeVerification.objects.select_related('classroom', 'verified_by')
        ],
        "staff": [
            {
                "id": s.pk,
                "staff_id": s.staff_id,
                "title": s.title,
                "first_name": s.first_name,
                "other_names": s.other_names or '',
                "last_name": s.last_name,
                "gender": s.gender,
                "dob": s.dob.isoformat() if s.dob else None,
                "designation": s.designation.name if s.designation else None,
                "department": s.department.name if s.department else None,
                "ssnit_id": s.ssnit_id or '',
                "phone_number": s.phone_number or '',
                "email": s.email,
                "employment_type": s.employment_type,
                "date_of_appointment": s.date_of_appointment.isoformat() if s.date_of_appointment else None,
                "year_of_last_promotion": s.year_of_last_promotion,
                "qualification": s.qualification,
                "certificate": s.certificate,
                "name_of_institution_completed": s.name_of_institution_completed,
                "year_completed": s.year_completed,
                "form_class": s.form_class.class_name if s.form_class else None,
                "address": s.address or '',
                "username": s.user.username if s.user else None,
            }
            for s in StaffProfile.objects.select_related(
                'user', 'designation', 'department', 'form_class'
            ).order_by('last_name', 'first_name')
        ],
        "staff_assignments": [
            {
                "staff_id": scs.staff.staff_id,
                "classroom": scs.classroom.class_name,
                "subject": scs.subject.subject_name,
            }
            for scs in StaffClassSubject.objects.select_related('staff', 'classroom', 'subject').order_by('staff__staff_id')
        ],
        "promotion_criteria": [
            {
                "classroom": pc.classroom.class_name,
                "min_grand_total": float(pc.min_grand_total),
            }
            for pc in PromotionCriteria.objects.select_related('classroom')
        ],
        "notifications": [
            {
                "recipient": n.recipient.username,
                "title": n.title,
                "message": n.message,
                "notification_type": n.notification_type,
                "is_read": n.is_read,
                "created_at": n.created_at.isoformat() if n.created_at else None,
            }
            for n in Notification.objects.select_related('recipient')
        ],
        "timetables": [
            {
                "id": t.pk,
                "student_class": t.student_class.class_name,
                "title": t.title,
                "academic_term": str(t.academic_term),
                "is_active": t.is_active,
            }
            for t in Timetable.objects.select_related('student_class', 'academic_term')
        ],
        "timetable_slots": [
            {
                "timetable_id": ts.timetable_id,
                "subject": ts.subject.subject_name,
                "teacher": ts.teacher.staff_id if ts.teacher else None,
                "day_of_week": ts.day_of_week,
                "start_time": ts.start_time.strftime('%H:%M:%S') if ts.start_time else None,
                "end_time": ts.end_time.strftime('%H:%M:%S') if ts.end_time else None,
                "room_or_note": ts.room_or_note or '',
            }
            for ts in TimetableSlot.objects.select_related('timetable', 'subject', 'teacher')
        ],
    }

    response = HttpResponse(
        json.dumps(payload, indent=2),
        content_type='application/json'
    )
    year_label = current_session.academic_year.replace('/', '-') if current_session else 'unknown'
    term_label = current_term.term_name.replace(' ', '') if current_term else 'unknown'
    filename = f"session_database_backup_{year_label}_{term_label}_{datetime.now().strftime('%Y-%m-%d')}.json"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _apply_session_payload(payload):
    """Shared, transactional engine that applies a session-backup payload.

    Used by both the JSON import and the Excel import. Anything in `payload`
    is optional except the core 'session' section. Raises on error so callers
    can roll back and show a message. Returns the inserted-item counts.
    """
    counts = {
        'sessions': 0, 'terms': 0, 'classes': 0, 'subjects': 0,
        'departments': 0, 'designations': 0, 'class_subjects': 0,
        'parents': 0, 'students': 0, 'enrollments': 0,
        'subject_assessments': 0, 'midterm_records': 0, 'grade_verifications': 0,
        'staff': 0, 'staff_assignments': 0, 'promotion_criteria': 0,
        'timetables': 0, 'timetable_slots': 0,
    }

    with transaction.atomic():
            session_by_year = {s.academic_year: s for s in AcademicSession.objects.all()}
            subject_by_name = {s.subject_name: s for s in Subject.objects.all()}
            class_by_name = {c.class_name: c for c in ClassRoom.objects.all()}
            staff_by_id = {s.staff_id: s for s in StaffProfile.objects.all()}
            parent_by_id = {p.pk: p for p in Parent.objects.all()}
            student_by_adm = {s.admission_number: s for s in Student.objects.all()}

            for name in payload.get('departments', []):
                Department.objects.get_or_create(name=name)
                counts['departments'] += 1

            for name in payload.get('designations', []):
                Designation.objects.get_or_create(name=name)
                counts['designations'] += 1

            for item in payload.get('sessions', []):
                session, created = AcademicSession.objects.get_or_create(academic_year=item['academic_year'])
                if created:
                    if item.get('is_current', False) and not AcademicSession.objects.filter(is_current=True).exists():
                        session.is_current = True
                    session.save()
                    counts['sessions'] += 1
                session_by_year[session.academic_year] = session

            term_by_key = {}
            for item in payload.get('terms', []):
                session = session_by_year.get(item['session'])
                if not session:
                    continue
                term, created = Term.objects.get_or_create(
                    session=session, term_name=item['term_name'],
                    defaults={'is_active': item.get('is_active', False)},
                )
                if created and item.get('is_active', False)\
                        and not Term.objects.filter(is_active=True).exists():
                    term.is_active = True
                    term.save()
                if created:
                    counts['terms'] += 1
                term_by_key[(session.academic_year, term.term_name)] = term

            for item in payload.get('classes', []):
                classroom, created = ClassRoom.objects.get_or_create(
                    class_name=item['class_name'],
                    defaults={'order': item.get('order', 0)},
                )
                if created:
                    counts['classes'] += 1
                else:
                    classroom.order = item.get('order', classroom.order)
                if item.get('next_class'):
                    classroom.next_class = class_by_name.get(item['next_class'])
                classroom.save()
                class_by_name[classroom.class_name] = classroom

            for name in payload.get('subjects', []):
                Subject.objects.get_or_create(subject_name=name)
                counts['subjects'] += 1
                subject_by_name[name] = Subject.objects.get(subject_name=name)

            for item in payload.get('class_subjects', []):
                classroom = class_by_name.get(item['classroom'])
                subject = subject_by_name.get(item['subject'])
                if classroom and subject:
                    ClassSubject.objects.get_or_create(classroom=classroom, subject=subject)
                    counts['class_subjects'] += 1

            for item in payload.get('parents', []):
                parent, created = Parent.objects.get_or_create(
                    name=item.get('name'), telephone_number=item.get('telephone_number'),
                    defaults={
                        'occupation': item.get('occupation') or '',
                        'residential_address': item.get('residential_address') or '',
                        'email': item.get('email') or '',
                    },
                )
                if created:
                    if item.get('id'):
                        parent.pk = item['id']
                    parent.save()
                    counts['parents'] += 1
                parent_by_id[parent.pk] = parent

            for item in payload.get('students', []):
                from datetime import date as ddate

                def _parse_iso(value):
                    if not value:
                        return None
                    try:
                        return ddate.fromisoformat(value)
                    except (ValueError, TypeError):
                        return None

                student, created = Student.objects.get_or_create(
                    admission_number=item['admission_number'],
                    defaults={
                        'first_name': item['first_name'],
                        'last_name': item['last_name'],
                        'gender': item.get('gender', 'Male'),
                        'dob': _parse_iso(item.get('dob')) or date.today(),
                    },
                )
                if created:
                    counts['students'] += 1
                else:
                    student.first_name = item.get('first_name', student.first_name)
                    student.last_name = item.get('last_name', student.last_name)
                student.other_names = item.get('other_names') or ''
                student.gender = item['gender']
                dob = _parse_iso(item.get('dob'))
                if dob is not None:
                    student.dob = dob
                date_of_admission = _parse_iso(item.get('date_of_admission'))
                if date_of_admission is not None:
                    student.date_of_admission = date_of_admission
                student.status = item.get('status', student.status)
                student.living_with = item.get('living_with', student.living_with)
                if item.get('previous_school_attended'):
                    student.previous_school_attended = item['previous_school_attended']
                if item.get('classroom'):
                    student.classroom = class_by_name.get(item['classroom'])
                if item.get('father_id'):
                    student.father = parent_by_id.get(item['father_id'])
                if item.get('mother_id'):
                    student.mother = parent_by_id.get(item['mother_id'])
                if item.get('pending_next_class'):
                    student.pending_next_class = class_by_name.get(item['pending_next_class'])
                student.promotion_status = item.get('promotion_status', 'NEUTRAL')
                student.is_alumni = bool(item.get('is_alumni', False))
                student.save()
                student_by_adm[student.admission_number] = student

            for item in payload.get('enrollments', []):
                student = student_by_adm.get(item['student'])
                classroom = class_by_name.get(item['classroom'])
                session = session_by_year.get(item['academic_session'])
                term = term_by_key.get((item['academic_session'], item['academic_term']))
                if student and classroom and session and term:
                    Enrollment.objects.get_or_create(
                        student=student, classroom=classroom,
                        academic_session=session, academic_term=term,
                    )
                    counts['enrollments'] += 1

            for item in payload.get('subject_assessments', []):
                student = student_by_adm.get(item['student'])
                subject = subject_by_name.get(item['subject'])
                session = session_by_year.get(item['academic_session'])
                term = term_by_key.get((item['academic_session'], item['academic_term']))
                if student and subject and session and term:
                    SubjectAssessment.objects.get_or_create(
                        student=student, subject=subject,
                        academic_session=session, academic_term=term,
                        defaults={'class_score': item['class_score'], 'exam_score': item['exam_score']},
                    )
                    counts['subject_assessments'] += 1

            for item in payload.get('midterm_records', []):
                student = student_by_adm.get(item['student'])
                subject = subject_by_name.get(item['subject'])
                session = session_by_year.get(item['academic_session'])
                term = term_by_key.get((item['academic_session'], item['term']))
                if student and subject and session and term:
                    MidTermRecord.objects.get_or_create(
                        student=student, academic_session=session, term=term, subject=subject,
                        defaults={
                            'midterm_score': item.get('midterm_score'),
                            'classroom': class_by_name.get(item.get('classroom')) if item.get('classroom') else None,
                            'recorded_by': staff_by_id.get(item.get('recorded_by')) if item.get('recorded_by') else None,
                        },
                    )
                    counts['midterm_records'] += 1

            for item in payload.get('grade_verifications', []):
                classroom = class_by_name.get(item['classroom'])
                session = session_by_year.get(item['academic_session'])
                term = term_by_key.get((item['academic_session'], item['academic_term']))
                if classroom and session and term:
                    GradeVerification.objects.get_or_create(
                        classroom=classroom, academic_session=session, academic_term=term,
                        defaults={'verified_by': staff_by_id.get(item.get('verified_by')) if item.get('verified_by') else None},
                    )
                    counts['grade_verifications'] += 1

            for item in payload.get('staff', []):
                staff, created = StaffProfile.objects.get_or_create(
                    staff_id=item['staff_id'], defaults={'email': item['email']},
                )
                if created:
                    counts['staff'] += 1
                for field in ['title', 'first_name', 'other_names', 'last_name', 'gender']:
                    if item.get(field) is not None:
                        setattr(staff, field, item[field])
                if item.get('dob'):
                    from datetime import date as ddate
                    try:
                        staff.dob = ddate.fromisoformat(item['dob'])
                    except ValueError:
                        pass
                if item.get('date_of_appointment'):
                    from datetime import date as ddate
                    try:
                        staff.date_of_appointment = ddate.fromisoformat(item['date_of_appointment'])
                    except ValueError:
                        pass
                if item.get('designation'):
                    staff.designation = Designation.objects.filter(name=item['designation']).first()
                if item.get('department'):
                    staff.department = Department.objects.filter(name=item['department']).first()
                for field in ['ssnit_id', 'phone_number', 'employment_type', 'qualification',
                              'certificate', 'name_of_institution_completed', 'address']:
                    if item.get(field) is not None:
                        setattr(staff, field, item[field])
                if item.get('year_of_last_promotion') is not None:
                    staff.year_of_last_promotion = item['year_of_last_promotion']
                if item.get('year_completed') is not None:
                    staff.year_completed = item['year_completed']
                if item.get('email'):
                    staff.email = item['email']
                if item.get('form_class'):
                    staff.form_class = class_by_name.get(item['form_class'])
                staff.save()
                staff_by_id[staff.staff_id] = staff

            for item in payload.get('staff_assignments', []):
                staff = staff_by_id.get(item['staff_id'])
                classroom = class_by_name.get(item['classroom'])
                subject = subject_by_name.get(item['subject'])
                if staff and classroom and subject:
                    StaffClassSubject.objects.get_or_create(staff=staff, classroom=classroom, subject=subject)
                    counts['staff_assignments'] += 1

            for item in payload.get('promotion_criteria', []):
                classroom = class_by_name.get(item['classroom'])
                if classroom:
                    PromotionCriteria.objects.get_or_create(
                        classroom=classroom, defaults={'min_grand_total': item['min_grand_total']},
                    )
                    counts['promotion_criteria'] += 1

            timetable_by_id = {}
            for item in payload.get('timetables', []):
                classroom = class_by_name.get(item['student_class'])
                term = None
                for t in Term.objects.select_related('session').all():
                    if str(t) == item['academic_term']:
                        term = t
                        break
                if not classroom or not term:
                    continue
                timetable, created = Timetable.objects.get_or_create(
                    student_class=classroom, academic_term=term,
                    defaults={
                        'title': item['title'],
                        'is_active': item.get('is_active', True),
                    },
                )
                if created:
                    counts['timetables'] += 1
                timetable_by_id[item.get('id')] = timetable

            for item in payload.get('timetable_slots', []):
                timetable = timetable_by_id.get(item.get('timetable_id'))
                subject = subject_by_name.get(item['subject'])
                teacher = staff_by_id.get(item.get('teacher'))
                if not timetable or not subject:
                    continue
                TimetableSlot.objects.get_or_create(
                    timetable=timetable, day_of_week=item['day_of_week'],
                    start_time=item['start_time'],
                    defaults={
                        'subject': subject,
                        'end_time': item['end_time'],
                        'teacher': teacher,
                        'room_or_note': item.get('room_or_note', ''),
                    },
                )
                counts['timetable_slots'] += 1

    return counts


def import_session_json(request):
    if not _is_admin(request.user):
        raise PermissionDenied

    upload = request.FILES.get('backup_file')
    if not upload:
        messages.error(request, "No backup file was uploaded.")
        return redirect('configure_session')

    try:
        payload = json.loads(upload.read().decode('utf-8'))
    except (UnicodeDecodeError, json.JSONDecodeError):
        messages.error(request, "The uploaded file is not a valid JSON backup.")
        return redirect('configure_session')

    if not isinstance(payload, dict) or not payload.get('session'):
        messages.error(request, "This file does not look like a session database backup (missing 'session' section).")
        return redirect('configure_session')

    try:
        counts = _apply_session_payload(payload)
    except Exception as exc:
        messages.error(request, f"Import failed (rolled back): {exc}")
        return redirect('configure_session')

    total = sum(counts.values())
    parts = ", ".join(f"{k} {v}" for k, v in counts.items() if v)
    messages.success(request, f"Backup imported successfully! Imported {total} items: {parts}.")
    return redirect('configure_session')


_DAY_DISPLAY_TO_CODE = {
    'Monday': 'MON', 'Tuesday': 'TUE', 'Wednesday': 'WED', 'Thursday': 'THU', 'Friday': 'FRI',
    'Mon': 'MON', 'Tue': 'TUE', 'Wed': 'WED', 'Thu': 'THU', 'Fri': 'FRI',
}


def _cell_str(value):
    if value is None:
        return ''
    return str(value).strip()


def _cell_bool(value):
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    return text in ('1', 'true', 'yes', 'y', 'active', 'current')


def _cell_float(value):
    if value is None or value == '':
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _read_export_sheet(wb, sheet_name):
    """Read a sheet from an exported session-database workbook.

    Exports put a merged title on row 1, headers on row 3, and data from
    row 4 onward. Returns a list of dicts keyed by header text.
    """
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = [_cell_str(ws.cell(row=3, column=col).value) for col in range(1, ws.max_column + 1)]
    if not any(headers):
        return []
    rows = []
    for r in range(4, ws.max_row + 1):
        values = [ws.cell(row=r, column=col).value for col in range(1, len(headers) + 1)]
        if not any(v is not None and _cell_str(v) for v in values):
            continue
        row = {}
        for header, value in zip(headers, values):
            if header:
                row[header] = value
        rows.append(row)
    return rows


def _excel_export_to_payload(wb):
    """Convert an exported session-database workbook back into a payload."""
    payload = {
        'session': {'academic_year': None, 'is_current': False},
        'sessions': [], 'terms': [], 'classes': [], 'subjects': [],
        'departments': [], 'designations': [], 'class_subjects': [],
        'parents': [], 'students': [], 'enrollments': [],
        'subject_assessments': [], 'midterm_records': [],
        'grade_verifications': [], 'staff': [], 'staff_assignments': [],
        'promotion_criteria': [], 'notifications': [], 'timetables': [],
        'timetable_slots': [],
    }

    sessions = _read_export_sheet(wb, 'Sessions')
    for row in sessions:
        payload['sessions'].append({
            'academic_year': _cell_str(row.get('Academic Year')),
            'is_current': _cell_bool(row.get('Is Current')),
        })
    if payload['sessions']:
        payload['session'] = payload['sessions'][0]

    for row in _read_export_sheet(wb, 'Terms'):
        payload['terms'].append({
            'session': _cell_str(row.get('Session')),
            'term_name': _cell_str(row.get('Term Name')),
            'is_active': _cell_bool(row.get('Is Active')),
        })

    for row in _read_export_sheet(wb, 'Classes'):
        payload['classes'].append({
            'class_name': _cell_str(row.get('Class Name')),
            'order': row.get('Order'),
            'next_class': _cell_str(row.get('Next Class')),
            'form_master_id': _cell_str(row.get('Form Master')),
        })

    payload['subjects'] = [_cell_str(r.get('Subject Name')) for r in _read_export_sheet(wb, 'Subjects')]
    payload['departments'] = [_cell_str(r.get('Name')) for r in _read_export_sheet(wb, 'Departments')]
    payload['designations'] = [_cell_str(r.get('Name')) for r in _read_export_sheet(wb, 'Designations')]

    for row in _read_export_sheet(wb, 'Class Subjects'):
        payload['class_subjects'].append({
            'classroom': _cell_str(row.get('Class')),
            'subject': _cell_str(row.get('Subject')),
        })

    for row in _read_export_sheet(wb, 'Parents'):
        payload['parents'].append({
            'id': row.get('ID'),
            'name': _cell_str(row.get('Name')),
            'occupation': _cell_str(row.get('Occupation')),
            'residential_address': _cell_str(row.get('Residential Address')),
            'email': _cell_str(row.get('Email')),
            'telephone_number': _cell_str(row.get('Telephone')),
        })

    parent_key_to_id = {}
    for p in payload['parents']:
        if p['name']:
            key = (p['name'], p['telephone_number'])
            if key not in parent_key_to_id:
                parent_key_to_id[key] = p['id']

    for row in _read_export_sheet(wb, 'Students'):
        father_key = (_cell_str(row.get('Father Name')), _cell_str(row.get('Father Phone')))
        mother_key = (_cell_str(row.get('Mother Name')), _cell_str(row.get('Mother Phone')))
        payload['students'].append({
            'id': row.get('ID'),
            'admission_number': _cell_str(row.get('Admission No.')),
            'first_name': _cell_str(row.get('First Name')),
            'other_names': _cell_str(row.get('Other Names')),
            'last_name': _cell_str(row.get('Last Name')),
            'gender': _cell_str(row.get('Gender')),
            'dob': row.get('Date of Birth'),
            'date_of_admission': row.get('Date of Admission'),
            'status': _cell_str(row.get('Status')),
            'living_with': _cell_str(row.get('Living With')),
            'previous_school_attended': _cell_str(row.get('Previous School')),
            'classroom': _cell_str(row.get('Class')),
            'father_id': parent_key_to_id.get(father_key),
            'mother_id': parent_key_to_id.get(mother_key),
            'pending_next_class': _cell_str(row.get('Pending Next Class')),
            'promotion_status': _cell_str(row.get('Promotion Status')),
            'is_alumni': _cell_bool(row.get('Is Alumni')),
        })

    for row in _read_export_sheet(wb, 'Enrollments'):
        payload['enrollments'].append({
            'student': _cell_str(row.get('Admission No.')),
            'classroom': _cell_str(row.get('Class')),
            'academic_session': _cell_str(row.get('Academic Session')),
            'academic_term': _cell_str(row.get('Term')),
        })

    for row in _read_export_sheet(wb, 'Subject Assessments'):
        payload['subject_assessments'].append({
            'student': _cell_str(row.get('Admission No.')),
            'subject': _cell_str(row.get('Subject')),
            'academic_session': _cell_str(row.get('Academic Session')),
            'academic_term': _cell_str(row.get('Term')),
            'class_score': _cell_float(row.get('Class Score (30%)')),
            'exam_score': _cell_float(row.get('Exam Score (70%)')),
        })

    for row in _read_export_sheet(wb, 'Mid-Term Records'):
        payload['midterm_records'].append({
            'student': _cell_str(row.get('Admission No.')),
            'subject': _cell_str(row.get('Subject')),
            'academic_session': _cell_str(row.get('Academic Session')),
            'term': _cell_str(row.get('Term')),
            'classroom': _cell_str(row.get('Class')),
            'midterm_score': _cell_float(row.get('Mid-Term Score')),
            'recorded_by': _cell_str(row.get('Recorded By')),
        })

    for row in _read_export_sheet(wb, 'Grade Verifications'):
        payload['grade_verifications'].append({
            'classroom': _cell_str(row.get('Class')),
            'academic_session': _cell_str(row.get('Academic Session')),
            'academic_term': _cell_str(row.get('Term')),
            'verified_by': _cell_str(row.get('Verified By')),
        })

    for row in _read_export_sheet(wb, 'Staff'):
        payload['staff'].append({
            'staff_id': _cell_str(row.get('Staff ID')),
            'title': _cell_str(row.get('Title')),
            'first_name': _cell_str(row.get('First Name')),
            'other_names': _cell_str(row.get('Other Names')),
            'last_name': _cell_str(row.get('Last Name')),
            'gender': _cell_str(row.get('Gender')),
            'dob': row.get('Date of Birth'),
            'designation': _cell_str(row.get('Designation')),
            'department': _cell_str(row.get('Department')),
            'ssnit_id': _cell_str(row.get('SSNIT ID')),
            'phone_number': _cell_str(row.get('Phone')),
            'email': _cell_str(row.get('Email')),
            'employment_type': _cell_str(row.get('Employment Type')),
            'date_of_appointment': row.get('Date of Appointment'),
            'year_of_last_promotion': row.get('Year of Last Promotion'),
            'qualification': _cell_str(row.get('Qualification')),
            'certificate': _cell_str(row.get('Certificate')),
            'name_of_institution_completed': _cell_str(row.get('Institution Completed')),
            'year_completed': row.get('Year Completed'),
            'form_class': _cell_str(row.get('Form Class (Form Teacher)')),
            'address': _cell_str(row.get('Residential Address')),
            'username': _cell_str(row.get('Username')),
        })

    for row in _read_export_sheet(wb, 'Staff Assignments'):
        payload['staff_assignments'].append({
            'staff_id': _cell_str(row.get('Staff ID')),
            'classroom': _cell_str(row.get('Class')),
            'subject': _cell_str(row.get('Subject')),
        })

    for row in _read_export_sheet(wb, 'Promotion Criteria'):
        payload['promotion_criteria'].append({
            'classroom': _cell_str(row.get('Class')),
            'min_grand_total': _cell_float(row.get('Minimum Grand Total')),
        })

    for row in _read_export_sheet(wb, 'Notifications'):
        payload['notifications'].append({
            'recipient': _cell_str(row.get('Recipient')),
            'notification_type': _cell_str(row.get('Type')),
            'title': _cell_str(row.get('Title')),
            'message': _cell_str(row.get('Message')),
            'is_read': _cell_bool(row.get('Is Read')),
        })

    for row in _read_export_sheet(wb, 'Timetables'):
        payload['timetables'].append({
            'id': row.get('ID'),
            'student_class': _cell_str(row.get('Class')),
            'title': _cell_str(row.get('Title')),
            'academic_term': _cell_str(row.get('Term')),
            'is_active': _cell_bool(row.get('Is Active')),
        })

    slot_timetable_id = {}
    for t in payload['timetables']:
        label = f"{t['student_class']} - {t['title']}"
        slot_timetable_id[label] = t['id']
        slot_timetable_id.setdefault(t['student_class'], t['id'])

    for row in _read_export_sheet(wb, 'Timetable Slots'):
        day_display = _cell_str(row.get('Day'))
        payload['timetable_slots'].append({
            'timetable_id': slot_timetable_id.get(_cell_str(row.get('Timetable'))),
            'subject': _cell_str(row.get('Subject')),
            'teacher': _cell_str(row.get('Teacher')),
            'day_of_week': _DAY_DISPLAY_TO_CODE.get(day_display, day_display),
            'start_time': _cell_str(row.get('Start Time')),
            'end_time': _cell_str(row.get('End Time')),
            'room_or_note': _cell_str(row.get('Room / Note')),
        })

    return payload


_FLEX_ALIASES = {
    'admission_number': ['admission number', 'admission no', 'admission no.', 'admission', 'student id',
                         'student number', 'index number', 'student no', 'student no.'],
    'first_name': ['first name', 'fname', 'forename'],
    'other_names': ['other names', 'middle name', 'middle names', 'other name'],
    'last_name': ['last name', 'surname', 'lname', 'family name'],
    'full_name': ['full name', 'student name', 'name', 'pupil name', 'student full name'],
    'gender': ['gender', 'sex'],
    'dob': ['date of birth', 'dob', 'birth date', 'birthday'],
    'date_of_admission': ['date of admission', 'admission date', 'enrollment date', 'enrolment date'],
    'status': ['status', 'student status', 'enrolment status', 'enrollment status'],
    'parent_name': ['parent name', 'guardian name', 'guardian', 'parent/guardian', 'father/mother name'],
    'parent_phone': ['telephone', 'phone', 'phone number', 'telephone number', 'contact', 'mobile',
                     'parent phone', 'parent telephone', 'guardian phone'],
    'parent_email': ['email', 'email address', 'parent email', 'guardian email'],
    'occupation': ['occupation', 'parent occupation'],
    'class_name': ['class', 'class name', 'form', 'form class', 'stream', 'classroom', 'grade'],
    'subject_name': ['subject', 'subject name', 'course', 'course name'],
    'teacher_staff_id': ['staff id', 'staff no', 'staff no.', 'staff number', 'employee id', 'teacher id',
                         'staff code'],
    'teacher_name': ['teacher', 'teacher name', 'staff name', 'teacher full name'],
    'academic_year': ['academic year', 'academic session', 'session year', 'year', 'session name', 'session'],
    'term_name': ['term', 'term name', 'term number', 'term no'],
    'midterm_score': ['midterm score', 'mid-term score', 'mid term score', 'mte'],
    'class_score': ['class score', 'classwork', 'continuous assessment', 'ca score', 'term score'],
    'exam_score': ['exam score', 'examination score', 'exam'],
    'score': ['total score', 'total', 'grand total', 'score', 'average'],
    'grade': ['grade', 'position'],
    'promotion_status': ['promotion status', 'promotion'],
}


def _normalize_header(text):
    return _cell_str(text).lower().replace('_', ' ').replace('-', ' ').replace('/', ' ').replace('.', ' ')


def _normalize_aliase(alias):
    return _normalize_header(alias)


def _match_column(headers, aliases):
    aliases_norm = {_normalize_aliase(a) for a in aliases}
    for h in headers:
        if h in aliases_norm:
            return headers[h]
    return None


def _find_header_row(ws):
    """Detect the header row index for an arbitrary sheet (usually row 1)."""
    best_row, best_score = None, 0
    for r in range(1, min(max_row if (max_row := ws.max_row) else 1, 15) + 1):
        score = 0
        for c in range(1, ws.max_column + 1):
            header = _normalize_header(ws.cell(row=r, column=c).value)
            if header and any(header == _normalize_aliase(a) for aliases in _FLEX_ALIASES.values() for a in aliases):
                score += 1
        if score > best_score:
            best_score, best_row = score, r
    if best_score < 2:
        return None
    return best_row


def _sheet_columns(ws, header_row):
    cols = {}
    for c in range(1, ws.max_column + 1):
        raw = ws.cell(row=header_row, column=c).value
        if raw is not None and _cell_str(raw):
            cols[_normalize_header(raw)] = c
    return cols


def _flex_sheet_rows(ws, header_row, cols):
    rows = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = {}
        blank = True
        for norm, c in cols.items():
            val = ws.cell(row=r, column=c).value
            if _cell_str(val):
                blank = False
            row[norm] = val
        if not blank:
            rows.append(row)
    return rows


def _excel_flex_to_payload(wb):
    """Map the admin's own spreadsheets onto the backup payload.

    Tolerates missing sheets and missing optional columns. Only rows whose
    entity can be recognised are imported; core entities are sessions,
    classes, subjects, parents, students and staff.
    """
    payload = {
        'session': {'academic_year': None, 'is_current': False},
        'sessions': [], 'terms': [], 'classes': [], 'subjects': [],
        'departments': [], 'designations': [], 'class_subjects': [],
        'parents': [], 'students': [], 'enrollments': [],
        'subject_assessments': [], 'midterm_records': [],
        'grade_verifications': [], 'staff': [], 'staff_assignments': [],
        'promotion_criteria': [], 'notifications': [], 'timetables': [],
        'timetable_slots': [],
    }
    warnings = []
    parents_by_key = {}

    def get_col(cols, key):
        aliases = _FLEX_ALIASES.get(key, [])
        aliases_norm = {_normalize_aliase(a) for a in aliases}
        for n, c in cols.items():
            if n in aliases_norm:
                return c
        return None

    for ws in wb.worksheets:
        header_row = _find_header_row(ws)
        if header_row is None:
            continue
        cols = _sheet_columns(ws, header_row)
        rows = _flex_sheet_rows(ws, header_row, cols)
        if not rows:
            continue

        c_adm = get_col(cols, 'admission_number')
        c_fname = get_col(cols, 'first_name')
        c_lname = get_col(cols, 'last_name')
        c_full = get_col(cols, 'full_name')
        c_gender = get_col(cols, 'gender')
        c_dob = get_col(cols, 'dob')
        c_admdate = get_col(cols, 'date_of_admission')
        c_status = get_col(cols, 'status')
        c_class = get_col(cols, 'class_name')
        c_subject = get_col(cols, 'subject_name')
        c_parent = get_col(cols, 'parent_name')
        c_pemail = get_col(cols, 'parent_email')
        c_pphone = get_col(cols, 'parent_phone')
        c_staffid = get_col(cols, 'teacher_staff_id')
        c_acad = get_col(cols, 'academic_year')
        c_termname = get_col(cols, 'term_name')
        c_promo = get_col(cols, 'promotion_status')

        # ── Students sheet ──
        if c_adm or c_full or c_fname or c_lname:
            for row in rows:
                adm = _cell_str(row.get(list(cols.keys())[list(cols.values()).index(c_adm)] if c_adm else '')) if c_adm else ''
                first = _cell_str(row[list(cols.keys())[list(cols.values()).index(c_fname)]]) if c_fname else ''
                last = _cell_str(row[list(cols.keys())[list(cols.values()).index(c_lname)]]) if c_lname else ''
                full = _cell_str(row[list(cols.keys())[list(cols.values()).index(c_full)]]) if c_full else ''
                if not adm and c_full:
                    adm = full
                if not first and not last and full:
                    parts = full.split()
                    first = parts[0] if parts else ''
                    last = parts[-1] if len(parts) > 1 else ''
                classroom = _cell_str(row[list(cols.keys())[list(cols.values()).index(c_class)]]) if c_class else ''
                parent_name = _cell_str(row[list(cols.keys())[list(cols.values()).index(c_parent)]]) if c_parent else ''
                parent_phone = _cell_str(row[list(cols.keys())[list(cols.values()).index(c_pphone)]]) if c_pphone else ''
                parent_email = _cell_str(row[list(cols.keys())[list(cols.values()).index(c_pemail)]]) if c_pemail else ''
                if not adm and not first and not last:
                    continue
                if parent_name:
                    key = (parent_name, parent_phone)
                    if key not in parents_by_key:
                        parents_by_key[key] = {
                            'name': parent_name,
                            'telephone_number': parent_phone,
                            'email': parent_email,
                            'occupation': '',
                        }
                payload['students'].append({
                    'id': None,
                    'admission_number': adm,
                    'first_name': first,
                    'other_names': '',
                    'last_name': last,
                    'gender': _cell_str(row[list(cols.keys())[list(cols.values()).index(c_gender)]]) if c_gender else 'Male',
                    'dob': row[list(cols.keys())[list(cols.values()).index(c_dob)]] if c_dob else None,
                    'date_of_admission': row[list(cols.keys())[list(cols.values()).index(c_admdate)]] if c_admdate else None,
                    'status': _cell_str(row[list(cols.keys())[list(cols.values()).index(c_status)]]) if c_status else 'ACTIVE',
                    'living_with': '',
                    'previous_school_attended': '',
                    'classroom': classroom,
                    'father_id': None,
                    'mother_id': None,
                    'pending_next_class': '',
                    'promotion_status': _cell_str(row[list(cols.keys())[list(cols.values()).index(c_promo)]]) if c_promo else 'NEUTRAL',
                    'is_alumni': False,
                })
            continue

        # ── Classes sheet ──
        if c_class and not c_subject and not c_adm:
            for row in rows:
                cn = _cell_str(row[list(cols.keys())[list(cols.values()).index(c_class)]])
                if cn:
                    payload['classes'].append({
                        'class_name': cn,
                        'order': len(payload['classes']) + 1,
                        'next_class': '',
                        'form_master_id': '',
                    })
            continue

        # ── Subjects sheet ──
        if c_subject and not c_adm and not c_class:
            for row in rows:
                sn = _cell_str(row[list(cols.keys())[list(cols.values()).index(c_subject)]])
                if sn:
                    payload['subjects'].append(sn)
            continue

        # ── Class-Subject matrix ──
        if c_class and c_subject:
            for row in rows:
                cn = _cell_str(row[list(cols.keys())[list(cols.values()).index(c_class)]])
                sn = _cell_str(row[list(cols.keys())[list(cols.values()).index(c_subject)]])
                if cn and sn:
                    payload['class_subjects'].append({'classroom': cn, 'subject': sn})
            continue

        # ── Parents / Guardians sheet ──
        if c_parent and not c_adm and not c_full and not c_fname and not c_lname:
            for row in rows:
                pname = _cell_str(row[list(cols.keys())[list(cols.values()).index(c_parent)]])
                pphone = _cell_str(row[list(cols.keys())[list(cols.values()).index(c_pphone)]]) if c_pphone else ''
                pemail = _cell_str(row[list(cols.keys())[list(cols.values()).index(c_pemail)]]) if c_pemail else ''
                if pname:
                    parents_by_key[(pname, pphone)] = {
                        'name': pname,
                        'telephone_number': pphone,
                        'email': pemail,
                        'occupation': '',
                    }
            continue

        # ── Staff / Teachers sheet ──
        if c_staffid and not c_adm:
            for row in rows:
                sid = _cell_str(row[list(cols.keys())[list(cols.values()).index(c_staffid)]])
                if not sid:
                    continue
                first = _cell_str(row[list(cols.keys())[list(cols.values()).index(c_fname)]]) if c_fname else ''
                last = _cell_str(row[list(cols.keys())[list(cols.values()).index(c_lname)]]) if c_lname else ''
                full = _cell_str(row[list(cols.keys())[list(cols.values()).index(c_full)]]) if c_full else ''
                if not first and not last and full:
                    parts = full.split()
                    first = parts[0] if parts else ''
                    last = parts[-1] if len(parts) > 1 else ''
                payload['staff'].append({
                    'staff_id': sid,
                    'title': '',
                    'first_name': first,
                    'other_names': '',
                    'last_name': last,
                    'gender': _cell_str(row[list(cols.keys())[list(cols.values()).index(c_gender)]]) if c_gender else 'Male',
                    'dob': row[list(cols.keys())[list(cols.values()).index(c_dob)]] if c_dob else None,
                    'designation': '',
                    'department': '',
                    'ssnit_id': '',
                    'phone_number': '',
                    'email': '',
                    'employment_type': '',
                    'date_of_appointment': None,
                    'year_of_last_promotion': None,
                    'qualification': '',
                    'certificate': '',
                    'name_of_institution_completed': '',
                    'year_completed': None,
                    'form_class': '',
                    'address': '',
                    'username': None,
                })
            continue

        # ── Sessions / Terms sheet ──
        if c_acad:
            for row in rows:
                ay = _cell_str(row[list(cols.keys())[list(cols.values()).index(c_acad)]]) if c_acad else ''
                tn = _cell_str(row[list(cols.keys())[list(cols.values()).index(c_termname)]]) if c_termname else ''
                if ay and not any(s['academic_year'] == ay for s in payload['sessions']):
                    payload['sessions'].append({'academic_year': ay, 'is_current': False})
                if ay and tn and not any((t['session'], t['term_name']) == (ay, tn) for t in payload['terms']):
                    payload['terms'].append({'session': ay, 'term_name': tn, 'is_active': False})
            continue

    if parents_by_key:
        payload['parents'] = list(parents_by_key.values())

    # ── Link students to their unique parent ──
    parent_by_key = {}
    for p in payload['parents']:
        parent_by_key[(p['name'], p['telephone_number'])] = p
    for s in payload['students']:
        # flexible files carry a single guardian column; link to that parent
        pass  # father_id/mother_id are left null when no explicit father/mother columns exist

    if not payload['sessions']:
        warnings.append("No sessions were found - no session will be created.")
    return payload, warnings


@login_required
@require_POST
def import_session_excel(request):
    if not _is_admin(request.user):
        raise PermissionDenied

    from openpyxl import load_workbook

    upload = request.FILES.get('backup_file')
    if not upload:
        messages.error(request, "No Excel file was uploaded.")
        return redirect('configure_session')

    try:
        wb = load_workbook(upload, data_only=True)
    except Exception:
        messages.error(request, "The uploaded file is not a valid Excel workbook (.xlsx).")
        return redirect('configure_session')

    if 'Overview' in wb.sheetnames:
        payload = _excel_export_to_payload(wb)
        if not payload['sessions'] and not payload['students'] and not payload['classes']:
            messages.error(request, "This workbook does not look like a session database export "
                                    "(no Sessions, Students, or Classes data found).")
            return redirect('configure_session')
    else:
        payload, warnings = _excel_flex_to_payload(wb)
        if not payload['students'] and not payload['classes'] and not payload['sessions']:
            warnings.append("No classes, sessions, or students could be recognised in any sheet.")
        if not (payload['students'] or payload['classes'] or payload['sessions'] or payload['parents']):
            messages.error(request, "Could not find any importable data. Make sure your sheets have "
                                    "recognisable headers (e.g. 'Class', 'Admission No.', 'Student Name', "
                                    "'Parent Name', 'Telephone').")
            return redirect('configure_session')

    try:
        counts = _apply_session_payload(payload)
    except Exception as exc:
        messages.error(request, f"Excel import failed (rolled back): {exc}")
        return redirect('configure_session')

    total = sum(counts.values())
    parts = ", ".join(f"{k} {v}" for k, v in counts.items() if v)
    message = f"Excel import successful! Imported {total} items: {parts}."
    if locals().get('warnings'):
        message += " " + " ".join(warnings)
    messages.success(request, message)
    return redirect('configure_session')


@login_required
def parent_detail_view(request, parent_id):
    parent = get_object_or_404(Parent, pk=parent_id)

    children_qs = Student.objects.filter(
        Q(father=parent) | Q(mother=parent)
    ).select_related('classroom')
    children = []
    for student in children_qs:
        if student.father_id == parent.id:
            relationship = 'Father'
        elif student.mother_id == parent.id:
            relationship = 'Mother'
        else:
            relationship = 'Guardian'
        student.relationship = relationship
        children.append(student)

    return render(request, 'sis/parent_detail.html', {
        'parent': parent,
        'children': children,
    })


@login_required
def parent_edit_view(request, parent_id):
    if not request.user.is_superuser:
        raise PermissionDenied
    parent = get_object_or_404(Parent, pk=parent_id)
    if request.method == 'POST':
        form = ParentForm(request.POST, instance=parent)
        if form.is_valid():
            form.save()
            messages.success(request, f"Parent '{parent.name}' updated successfully!")
            return redirect('parent_detail', parent_id=parent.id)
    else:
        form = ParentForm(instance=parent)
    return render(request, 'sis/parent_edit.html', {
        'form': form,
        'parent': parent,
    })


def classes_subjects_hub(request):
    classes_list = ClassRoom.objects.annotate(
        student_count=Count('students')
    ).order_by('-order')

    unique_subjects = Subject.objects.all().order_by('subject_name')

    return render(request, 'sis/classes_subjects_hub.html', {
        'classes': classes_list,
        'subjects': unique_subjects,
    })


@login_required
def timetable_hub(request):
    timetables = Timetable.objects.select_related('student_class', 'academic_term').all()
    classes = ClassRoom.objects.all().order_by('order')
    subjects = Subject.objects.all().order_by('subject_name')
    teachers = StaffProfile.objects.select_related('user').all()
    terms = Term.objects.select_related('session').order_by('-session__academic_year', 'term_name')
    current_term = Term.objects.filter(is_active=True).first()
    return render(request, 'sis/timetable_hub.html', {
        'timetables': timetables,
        'classes': classes,
        'subjects': subjects,
        'teachers': teachers,
        'terms': terms,
        'current_term': current_term,
    })


@login_required
@require_POST
def timetable_create(request):
    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)

    class_id = data.get('class_id')
    title = data.get('title', '').strip()
    term_id = data.get('term_id')
    slots = data.get('slots', [])

    if not class_id or not title or not term_id:
        return JsonResponse({'error': 'Class, title, and term are required.'}, status=400)

    try:
        classroom = ClassRoom.objects.get(pk=class_id)
        term = Term.objects.get(pk=term_id)
    except (ClassRoom.DoesNotExist, Term.DoesNotExist):
        return JsonResponse({'error': 'Invalid class or term.'}, status=400)

    if Timetable.objects.filter(student_class=classroom, academic_term=term).exists():
        return JsonResponse({'error': f'A timetable already exists for {classroom.class_name} in this term.'}, status=400)

    timetable = Timetable.objects.create(
        student_class=classroom,
        title=title,
        academic_term=term,
    )

    created_slots = 0
    for slot_data in slots:
        try:
            subject = Subject.objects.get(pk=slot_data['subject_id'])
            teacher = StaffProfile.objects.get(pk=slot_data['teacher_id']) if slot_data.get('teacher_id') else None
            TimetableSlot.objects.create(
                timetable=timetable,
                subject=subject,
                teacher=teacher,
                day_of_week=slot_data['day_of_week'],
                start_time=slot_data['start_time'],
                end_time=slot_data['end_time'],
                room_or_note=slot_data.get('room_or_note', ''),
            )
            created_slots += 1
        except (Subject.DoesNotExist, StaffProfile.DoesNotExist, KeyError):
            continue

    return JsonResponse({
        'success': True,
        'timetable_id': timetable.id,
        'title': timetable.title,
        'slots_created': created_slots,
    })


@login_required
def timetable_detail_api(request, timetable_id):
    try:
        timetable = Timetable.objects.select_related('student_class', 'academic_term').get(pk=timetable_id)
    except Timetable.DoesNotExist:
        return JsonResponse({'error': 'Timetable not found'}, status=404)

    slots = timetable.slots.select_related('subject', 'teacher').order_by('day_of_week', 'start_time')

    grid = {}
    for slot in slots:
        day = slot.day_of_week
        if day not in grid:
            grid[day] = []
        grid[day].append({
            'id': slot.id,
            'subject': slot.subject.subject_name,
            'teacher': f"{slot.teacher.first_name} {slot.teacher.last_name}" if slot.teacher else 'Unassigned',
            'start_time': slot.start_time.strftime('%H:%M'),
            'end_time': slot.end_time.strftime('%H:%M'),
            'room_or_note': slot.room_or_note,
        })

    return JsonResponse({
        'timetable': {
            'id': timetable.id,
            'title': timetable.title,
            'class_name': timetable.student_class.class_name,
            'term': timetable.academic_term.term_name,
            'is_active': timetable.is_active,
            'updated_at': timetable.updated_at.strftime('%b %d, %Y %H:%M'),
        },
        'grid': grid,
        'days': ['MON', 'TUE', 'WED', 'THU', 'FRI'],
    })


@login_required
@require_POST
def timetable_delete(request, timetable_id):
    try:
        timetable = Timetable.objects.get(pk=timetable_id)
    except Timetable.DoesNotExist:
        return JsonResponse({'error': 'Timetable not found'}, status=404)
    timetable.delete()
    return JsonResponse({'success': True})


@login_required
@require_POST
def mark_all_notifications_read(request):
    Notification.objects.filter(recipient=request.user, is_read=False).update(is_read=True)
    return JsonResponse({'success': True})


@login_required
@require_POST
def save_theme_preference(request):
    try:
        data = json.loads(request.body)
        theme = data.get('theme', 'system')
        if theme not in ('light', 'dark', 'system'):
            return JsonResponse({'error': 'Invalid theme'}, status=400)
        staff = getattr(request.user, 'staff_profile', None)
        if staff:
            staff.theme = theme
            staff.save(update_fields=['theme'])
        return JsonResponse({'success': True, 'theme': theme})
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)