import io

from PIL import Image
from django.core.files.uploadedfile import SimpleUploadedFile
from datetime import date
from unittest.mock import patch

from django.test import RequestFactory, TestCase
from django.urls import reverse

from .forms import StaffRegistrationForm, StudentRegistrationForm
from .models import ClassRoom, Department, Designation, Parent, Student, Subject, SubjectAssessment, StaffProfile, AcademicSession, Term, ClassSubject, Enrollment


class StudentRegistrationTests(TestCase):
    def setUp(self):
        from django.contrib.auth.models import User
        self.user = User.objects.create_superuser(username="admin", password="password", email="admin@example.com")
        self.client.login(username="admin", password="password")

    def test_dashboard_calendar_partial_renders_without_full_page_shell(self):
        from django.contrib.auth.models import User

        user = User.objects.create_user(username="staff", password="password", email="staff@example.com")
        self.client.login(username="staff", password="password")

        response = self.client.get(reverse("dashboard"), {"month": 7, "year": 2026, "calendar_partial": 1})

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "sis/partials/dashboard_calendar.html")
        self.assertNotContains(response, "<!DOCTYPE html>")
        self.assertContains(response, "Calendar")

    def test_dashboard_calendar_marks_today_in_current_month(self):
        from .views import _build_dashboard_calendar_data

        request = RequestFactory().get(reverse("dashboard"))
        request.user = self.user

        class FixedDate(date):
            @classmethod
            def today(cls):
                return cls(2026, 7, 25)

        with patch('sis.views.date', FixedDate):
            calendar_data = _build_dashboard_calendar_data(request, date(2026, 7, 1))

        today_marked = [
            day for week in calendar_data['calendar_weeks'] for day in week if day.get('is_today')
        ]

        self.assertEqual(len(today_marked), 1)
        self.assertEqual(today_marked[0]['day'], 25)

    def test_e_path_redirects_to_students(self):
        response = self.client.get("/e")
        self.assertRedirects(response, reverse("dashboard"))

    def test_registration_page_renders(self):
        response = self.client.get(reverse("student_registration"))
        self.assertEqual(response.status_code, 200)

    def test_staff_registration_page_renders(self):
        response = self.client.get(reverse("register_staff"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Staff Registration")

    def test_staff_registration_saves_and_redirects(self):
        classroom = ClassRoom.objects.create(class_name="JHS 3")
        subject = Subject.objects.create(subject_name="ICT")
        response = self.client.post(reverse("register_staff"), {
            "title": "Mr.",
            "first_name": "Kwame",
            "last_name": "Boateng",
            "staff_id": "STAFF-001",
            "gender": "Male",
            "dob": "1985-01-01",
            "designation": "Teacher",
            "email": "kwame@example.com",
            "employment_type": "Permanent",
            "date_of_appointment": "2020-01-01",
            "department": "Mathematics",
            "qualification": "Degree",
            "certificate": "B.Ed",
            "name_of_institution_completed": "UCC",
            "year_completed": "2020",
            "form_class": classroom.id,
            "subject_areas": [subject.id],
        })
        self.assertEqual(response.status_code, 302)
        staff = StaffProfile.objects.get(staff_id="STAFF-001")
        self.assertEqual(f"{staff.first_name} {staff.last_name}", "Kwame Boateng")
        self.assertIn(subject, staff.subject_areas.all())

    def test_staff_registration_uploads_profile_picture(self):
        classroom = ClassRoom.objects.create(class_name="JHS 2")
        subject = Subject.objects.create(subject_name="Science")
        image_bytes = io.BytesIO()
        Image.new("RGB", (100, 100), color="blue").save(image_bytes, format="PNG")
        uploaded_file = SimpleUploadedFile(
            "avatar.png",
            image_bytes.getvalue(),
            content_type="image/png",
        )

        response = self.client.post(reverse("register_staff"), {
            "title": "Ms.",
            "first_name": "Ama",
            "last_name": "Boateng",
            "staff_id": "STAFF-002",
            "gender": "Female",
            "dob": "1990-02-02",
            "designation": "Teacher",
            "email": "ama@example.com",
            "employment_type": "Permanent",
            "date_of_appointment": "2021-01-01",
            "department": "Science",
            "qualification": "Degree",
            "certificate": "B.Ed",
            "name_of_institution_completed": "KNUST",
            "year_completed": "2021",
            "form_class": classroom.id,
            "subject_areas": [subject.id],
            "profile_picture": uploaded_file,
        })

        self.assertEqual(response.status_code, 302)
        staff = StaffProfile.objects.get(staff_id="STAFF-002")
        self.assertTrue(staff.profile_picture)

    def test_staff_avatar_initial_uses_name_when_no_picture(self):
        staff = StaffProfile.objects.create(
            first_name="Ada",
            last_name="Lovelace",
            staff_id="STAFF-003",
            email="ada@example.com",
        )

        self.assertEqual(staff.avatar_initial, "A")

    def test_student_list_renders_with_classrooms(self):
        classroom1 = ClassRoom.objects.create(class_name="JHS 1")
        classroom2 = ClassRoom.objects.create(class_name="JHS 2")
        response = self.client.get(reverse("student_list"))
        self.assertEqual(response.status_code, 200)
        self.assertIn("classrooms", response.context)
        self.assertEqual(len(response.context["classrooms"]), 2)
        self.assertContains(response, "JHS 1")
        self.assertContains(response, "JHS 2")

    def test_registration_form_has_gender_choices(self):
        form = StudentRegistrationForm()
        self.assertIn(("Male", "Male"), form.fields["gender"].choices)
        self.assertIn(("Female", "Female"), form.fields["gender"].choices)

    def test_class_report_page_renders(self):
        classroom = ClassRoom.objects.create(class_name="JHS 1")
        session = AcademicSession.objects.create(academic_year="2025/2026", is_current=True)
        term = Term.objects.create(session=session, term_name="Term 1", is_active=True)
        student = Student.objects.create(
            admission_number="001",
            first_name="Ada",
            last_name="Lovelace",
            gender="Female",
            dob="2000-01-01",
            status="Day",
            current_class=classroom,
        )
        subject = Subject.objects.create(subject_name="Mathematics")
        ClassSubject.objects.create(classroom=classroom, subject=subject)
        SubjectAssessment.objects.create(
            student=student,
            subject=subject,
            academic_session=session,
            academic_term=term,
            class_score=20,
            exam_score=30,
        )

        response = self.client.get(reverse("class_report_card_short", kwargs={"class_id": classroom.id}))
        self.assertEqual(response.status_code, 200)

    def test_parent_can_be_linked_to_student(self):
        classroom = ClassRoom.objects.create(class_name="JHS 2")
        father = Parent.objects.create(
            name="John Doe",
            occupation="Engineer",
            residential_address="Accra",
            email="john@example.com",
            telephone_number="0200000000",
        )
        mother = Parent.objects.create(
            name="Jane Doe",
            occupation="Teacher",
            residential_address="Accra",
            email="jane@example.com",
            telephone_number="0200000001",
        )
        student = Student.objects.create(
            admission_number="002",
            first_name="Ben",
            last_name="Doe",
            gender="Male",
            dob="2001-02-02",
            status="Boarder",
            living_with="Both",
            previous_school_attended="Old School",
            father=father,
            mother=mother,
            current_class=classroom,
        )

        self.assertEqual(student.father.name, "John Doe")
        self.assertEqual(student.mother.name, "Jane Doe")

    def test_parent_form_fields_optional(self):
        from .forms import ParentForm
        form = ParentForm()
        for field_name, field in form.fields.items():
            self.assertFalse(field.required, f"Field {field_name} should be optional")

    def test_register_student_no_parents(self):
        classroom = ClassRoom.objects.create(class_name="Class A")
        post_data = {
            "admission_number": "1001",
            "first_name": "Child",
            "last_name": "One",
            "dob": "2015-05-05",
            "gender": "Male",
            "status": "Day",
            "living_with": "Both",
            "previous_school_attended": "None",
            "current_class": classroom.id,
            # Father details (empty)
            "father-name": "",
            "father-occupation": "",
            "father-residential_address": "",
            "father-email": "",
            "father-telephone_number": "",
            # Mother details (empty)
            "mother-name": "",
            "mother-occupation": "",
            "mother-residential_address": "",
            "mother-email": "",
            "mother-telephone_number": "",
        }
        response = self.client.post(reverse("student_registration"), post_data)
        self.assertEqual(response.status_code, 302) # Redirects to student_list
        student = Student.objects.get(admission_number="1001")
        self.assertIsNone(student.father)
        self.assertIsNone(student.mother)

    def test_register_student_only_mother(self):
        classroom = ClassRoom.objects.create(class_name="Class B")
        post_data = {
            "admission_number": "1002",
            "first_name": "Child",
            "last_name": "Two",
            "dob": "2015-05-05",
            "gender": "Female",
            "status": "Day",
            "living_with": "Mother",
            "previous_school_attended": "None",
            "current_class": classroom.id,
            # Father details (empty)
            "father-name": "",
            "father-occupation": "",
            "father-residential_address": "",
            "father-email": "",
            "father-telephone_number": "",
            # Mother details (filled)
            "mother-name": "Jane Mother",
            "mother-occupation": "Engineer",
            "mother-residential_address": "Home Address",
            "mother-email": "jane@example.com",
            "mother-telephone_number": "0241234567",
        }
        response = self.client.post(reverse("student_registration"), post_data)
        self.assertEqual(response.status_code, 302)
        student = Student.objects.get(admission_number="1002")
        self.assertIsNone(student.father)
        self.assertIsNotNone(student.mother)
        self.assertEqual(student.mother.name, "Jane Mother")
        self.assertEqual(student.mother.telephone_number, "0241234567")

    def test_register_student_parent_missing_name_and_phone(self):
        classroom = ClassRoom.objects.create(class_name="Class C")
        post_data = {
            "admission_number": "1003",
            "first_name": "Child",
            "last_name": "Three",
            "dob": "2015-05-05",
            "gender": "Male",
            "status": "Day",
            "living_with": "Both",
            "previous_school_attended": "None",
            "current_class": classroom.id,
            # Father details (occupation only - no name/phone)
            "father-name": "",
            "father-occupation": "Doctor",
            "father-residential_address": "",
            "father-email": "",
            "father-telephone_number": "",
            # Mother details (empty)
            "mother-name": "",
            "mother-occupation": "",
            "mother-residential_address": "",
            "mother-email": "",
            "mother-telephone_number": "",
        }
        response = self.client.post(reverse("student_registration"), post_data)
        self.assertEqual(response.status_code, 302)
        student = Student.objects.get(admission_number="1003")
        self.assertIsNone(student.father)
        self.assertIsNone(student.mother)

    def test_register_student_parent_only_phone(self):
        classroom = ClassRoom.objects.create(class_name="Class D")
        post_data = {
            "admission_number": "1004",
            "first_name": "Child",
            "last_name": "Four",
            "dob": "2015-05-05",
            "gender": "Female",
            "status": "Day",
            "living_with": "Both",
            "previous_school_attended": "None",
            "current_class": classroom.id,
            # Father details (telephone only)
            "father-name": "",
            "father-occupation": "",
            "father-residential_address": "",
            "father-email": "",
            "father-telephone_number": "0509999999",
            # Mother details (empty)
            "mother-name": "",
            "mother-occupation": "",
            "mother-residential_address": "",
            "mother-email": "",
            "mother-telephone_number": "",
        }
        response = self.client.post(reverse("student_registration"), post_data)
        self.assertEqual(response.status_code, 302)
        student = Student.objects.get(admission_number="1004")
        self.assertIsNotNone(student.father)
        self.assertIsNone(student.mother)
        self.assertEqual(student.father.telephone_number, "0509999999")

    def test_bulk_grade_entry_saves_and_redirects_to_same_view(self):
        classroom = ClassRoom.objects.create(class_name="Class E")
        subject = Subject.objects.create(subject_name="Science")
        session = AcademicSession.objects.create(academic_year="2025/2026", is_current=True)
        term = Term.objects.create(session=session, term_name="Term 1", is_active=True)
        student = Student.objects.create(
            admission_number="1005",
            first_name="Alice",
            last_name="Test",
            dob="2010-01-01",
            gender="Female",
            status="Day",
            current_class=classroom,
        )
        post_data = {
            f"class_score_{student.id}": "25.5",
            f"exam_score_{student.id}": "45.0",
        }
        url = reverse("bulk_grade_entry", kwargs={"class_id": classroom.id, "subject_id": subject.id})
        response = self.client.post(url, post_data)
        
        # Verify it redirects to the same bulk grade entry view
        self.assertRedirects(response, url)
        
        # Verify assessment was saved correctly
        assessment = SubjectAssessment.objects.get(student=student, subject=subject)
        self.assertEqual(float(assessment.class_score), 25.5)
        self.assertEqual(float(assessment.exam_score), 45.0)

        # Verify messages contains the success message
        messages_list = list(response.wsgi_request._messages)
        self.assertEqual(len(messages_list), 1)
        self.assertEqual(str(messages_list[0]), f"Grades for Science saved successfully!")

    def test_class_report_empty_state_displays_banner(self):
        classroom = ClassRoom.objects.create(class_name="Class Empty")
        url = reverse("class_report", kwargs={"class_id": classroom.id})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertIn("classrooms", response.context)
        self.assertFalse(response.context["has_graded_records"])
        self.assertContains(response, "No student grade records compiled for this class yet")

    def test_class_report_empty_state_shows_master_records_pill(self):
        classroom = ClassRoom.objects.create(class_name="Class Empty Pill")
        url = reverse("class_report", kwargs={"class_id": classroom.id})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context["has_graded_records"])
        self.assertContains(response, "View Master Records")

    def test_admin_report_landing_shows_staff_empty_state_with_pill(self):
        from django.contrib.auth.models import User

        self.client.logout()
        User.objects.create_superuser(username="repadmin", password="password")
        self.client.login(username="repadmin", password="password")
        classroom = ClassRoom.objects.create(class_name="Class Admin Empty")
        url = reverse("class_report", kwargs={"class_id": classroom.id})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context["has_graded_records"])
        self.assertTrue(response.context["has_class_subject_assignment"])
        self.assertTrue(response.context["is_admin_landing"])
        self.assertContains(response, "Group%208381.svg")
        self.assertContains(response, "View Master Records")
        self.assertContains(response, "No student grade records compiled for this class yet")

    def test_admin_report_landing_with_data_still_shows_empty_design_then_master(self):
        from django.contrib.auth.models import User

        self.client.logout()
        User.objects.create_superuser(username="repadmin2", password="password")
        self.client.login(username="repadmin2", password="password")
        classroom = ClassRoom.objects.create(class_name="Class Admin Data")
        session = AcademicSession.objects.create(academic_year="2025/2026", is_current=True)
        term = Term.objects.create(session=session, term_name="Term 1", is_active=True)
        student = Student.objects.create(
            admission_number="ADM-1",
            first_name="Data",
            last_name="Kid",
            gender="Female",
            dob="2011-01-01",
            status="Day",
            current_class=classroom,
        )
        subject = Subject.objects.create(subject_name="Maths")
        ClassSubject.objects.create(classroom=classroom, subject=subject)
        SubjectAssessment.objects.create(
            student=student, subject=subject,
            academic_session=session, academic_term=term,
            class_score=20, exam_score=40,
        )
        url = reverse("class_report", kwargs={"class_id": classroom.id})

        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["has_graded_records"])
        self.assertTrue(response.context["is_admin_landing"])
        self.assertContains(response, "View Master Records")

        master = self.client.get(url + "?master=1")
        self.assertEqual(master.status_code, 200)
        self.assertFalse(master.context["is_admin_landing"])
        self.assertTrue(master.context["is_master"])
        self.assertContains(master, "Data Kid")

    def test_class_report_with_grades_displays_table(self):
        classroom = ClassRoom.objects.create(class_name="Class Graded")
        session = AcademicSession.objects.create(academic_year="2025/2026", is_current=True)
        term = Term.objects.create(session=session, term_name="Term 1", is_active=True)
        student = Student.objects.create(
            admission_number="1006",
            first_name="Evelyn",
            last_name="Standings",
            dob="2010-01-01",
            gender="Female",
            status="Day",
            current_class=classroom,
        )
        subject = Subject.objects.create(subject_name="English")
        ClassSubject.objects.create(classroom=classroom, subject=subject)
        SubjectAssessment.objects.create(
            student=student,
            subject=subject,
            academic_session=session,
            academic_term=term,
            class_score=30,
            exam_score=50,
        )
        url = reverse("class_report", kwargs={"class_id": classroom.id})
        response = self.client.get(url)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["has_graded_records"])
        self.assertTrue(response.context["is_admin_landing"])
        self.assertContains(response, "No student grade records compiled for this class yet")

        master = self.client.get(url + "?master=1")
        self.assertEqual(master.status_code, 200)
        self.assertTrue(master.context["has_graded_records"])
        self.assertFalse(master.context["is_admin_landing"])
        self.assertContains(master, "Evelyn Standings")



class FormMasterStudentVisibilityTests(TestCase):
    def setUp(self):
        from django.contrib.auth.models import User

        self.classroom = ClassRoom.objects.create(class_name="JHS 1")
        self.student = Student.objects.create(
            admission_number="FM-001",
            first_name="Kofi",
            last_name="Mensah",
            gender="Male",
            dob="2011-01-01",
            status="Day",
            classroom=self.classroom,
        )
        user = User.objects.create_user(username="formmaster", password="password")
        StaffProfile.objects.create(
            staff_id="FM-TEACHER-001",
            first_name="Form",
            last_name="Master",
            email="formmaster@example.com",
            user=user,
            form_class=self.classroom,
        )
        self.client.login(username="formmaster", password="password")

    def test_form_master_sees_students_of_form_class_in_directory(self):
        response = self.client.get(reverse("student_list"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Kofi  Mensah")

    def test_form_master_without_teaching_assignment_still_sees_form_class(self):
        other_class = ClassRoom.objects.create(class_name="JHS 2")
        other_student = Student.objects.create(
            admission_number="FM-002",
            first_name="Ama",
            last_name="Serwaa",
            gender="Female",
            dob="2011-02-02",
            status="Day",
            classroom=other_class,
        )
        response = self.client.get(reverse("student_list"))
        self.assertContains(response, "Kofi  Mensah")
        self.assertNotContains(response, "Ama  Serwaa")


class ExportStaffExcelTests(TestCase):
    def setUp(self):
        import io

        from django.contrib.auth.models import User

        self._io = io
        self.user = User.objects.create_superuser(username="staffexport", password="password")
        self.client.login(username="staffexport", password="password")

    def test_export_staff_excel_includes_staff_rows(self):
        from django.contrib.auth.models import User

        dept = Department.objects.create(name="Sciences")
        des = Designation.objects.create(name="Head of Maths")
        cls = ClassRoom.objects.create(class_name="JHS 2", order=2)
        subject = Subject.objects.create(subject_name="Physics")
        s_user = User.objects.create_user(username="t1", email="t1@example.com", password="x")
        staff = StaffProfile.objects.create(
            staff_id="STF-100",
            title="Mr.",
            first_name="Kwesi",
            other_names="Kofi",
            last_name="Antwi",
            gender="Male",
            email="kwesi@example.com",
            phone_number="0244001234",
            department=dept,
            designation=des,
            form_class=cls,
            user=s_user,
        )
        staff.subject_areas.add(subject)

        response = self.client.get(reverse("export_staff_excel"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("Staff_Directory", response["Content-Disposition"])

        from openpyxl import load_workbook

        ws = load_workbook(self._io.BytesIO(response.content)).active
        header_row = [c.value for c in ws[3]]
        self.assertIn("Staff ID", header_row)
        self.assertIn("Designation", header_row)
        self.assertIn("Department", header_row)
        self.assertIn("Form Class", header_row)
        self.assertIn("Subject Areas", header_row)

        cell_values = [c.value for row in ws.iter_rows(min_row=4) for c in row if c.value is not None]
        self.assertIn("STF-100", cell_values)
        self.assertIn("Kwesi", cell_values)
        self.assertIn("Antwi", cell_values)
        self.assertIn("Sciences", cell_values)
        self.assertIn("Physics", cell_values)

    def test_export_staff_fab_only_for_superuser(self):
        response = self.client.get(reverse("staff_list"))
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.context["can_export"])
        self.assertContains(response, "Export to Excel")

        self.client.logout()
        from django.contrib.auth.models import User

        User.objects.create_user(username="plainstaff", password="password")
        self.client.login(username="plainstaff", password="password")
        response = self.client.get(reverse("staff_list"))
        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context["can_export"])
        self.assertNotContains(response, "Export to Excel")

    def test_export_staff_excel_requires_superuser(self):
        self.client.logout()
        from django.contrib.auth.models import User

        User.objects.create_user(username="plainstaff2", password="password")
        self.client.login(username="plainstaff2", password="password")
        response = self.client.get(reverse("export_staff_excel"))
        self.assertEqual(response.status_code, 403)


class ExportExcelTests(TestCase):
    def setUp(self):
        import io

        from django.contrib.auth.models import User
        from PIL import Image

        self._io = io
        self.user = User.objects.create_superuser(username="exportadmin", password="password")
        self.client.login(username="exportadmin", password="password")

    def test_export_students_excel_includes_linked_parents(self):
        classroom = ClassRoom.objects.create(class_name="JHS 1")
        father = Parent.objects.create(name="Kwame Mensah", telephone_number="0244000001")
        mother = Parent.objects.create(name="Ama Mensah", email="ama@example.com")
        Student.objects.create(
            admission_number="EX-001",
            first_name="Kofi",
            last_name="Mensah",
            gender="Male",
            dob="2011-01-01",
            status="Day",
            classroom=classroom,
            father=father,
            mother=mother,
        )

        response = self.client.get(reverse("export_students_excel"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("Students_Directory", response["Content-Disposition"])

        from openpyxl import load_workbook

        ws = load_workbook(self._io.BytesIO(response.content)).active
        header_row = [c.value for c in ws[3]]
        self.assertIn("Father Name", header_row)
        self.assertIn("Mother Phone", header_row)

        cell_values = [c.value for row in ws.iter_rows(min_row=4) for c in row if c.value is not None]
        self.assertIn("EX-001", cell_values)
        self.assertIn("Kwame Mensah", cell_values)
        self.assertIn("ama@example.com", cell_values)

    def test_export_parents_excel_includes_linked_students(self):
        classroom = ClassRoom.objects.create(class_name="JHS 1")
        parent = Parent.objects.create(name="Kwame Mensah", telephone_number="0244000001")
        Student.objects.create(
            admission_number="EX-002",
            first_name="Kofi",
            last_name="Mensah",
            gender="Male",
            dob="2011-01-01",
            status="Day",
            classroom=classroom,
            father=parent,
        )

        response = self.client.get(reverse("export_parents_excel"))
        self.assertEqual(response.status_code, 200)
        self.assertIn("Parents_Directory", response["Content-Disposition"])

        from openpyxl import load_workbook

        ws = load_workbook(self._io.BytesIO(response.content)).active
        header_row = [c.value for c in ws[3]]
        self.assertIn("Parent Name", header_row)
        self.assertIn("Children (Name - Class)", header_row)

        cell_values = [c.value for row in ws.iter_rows(min_row=4) for c in row if c.value is not None]
        self.assertIn("Kwame Mensah", cell_values)
        self.assertTrue(any("Kofi Mensah - JHS 1" in str(v) for v in cell_values))


class ExportSessionTests(TestCase):
    def setUp(self):
        import io

        from django.contrib.auth.models import User

        self._io = io
        self.session = AcademicSession.objects.create(academic_year="2025/2026", is_current=True)
        self.term = Term.objects.create(session=self.session, term_name="Term 1", is_active=True)

        classroom = ClassRoom.objects.create(class_name="JHS 1")
        parent = Parent.objects.create(name="Kwame Mensah", telephone_number="0244000001")
        self.student = Student.objects.create(
            admission_number="EX-010",
            first_name="Kofi",
            other_names="Kojo",
            last_name="Mensah",
            gender="Male",
            dob="2011-01-01",
            status="Day",
            classroom=classroom,
            father=parent,
        )
        subject = Subject.objects.create(subject_name="Mathematics")
        SubjectAssessment.objects.create(
            student=self.student,
            subject=subject,
            academic_session=self.session,
            academic_term=self.term,
            class_score=40,
            exam_score=50,
        )

    def test_exports_require_superuser(self):
        from django.contrib.auth.models import User

        staff_user = User.objects.create_user(username="exportstaff", password="password")
        self.client.login(username="exportstaff", password="password")

        excel_resp = self.client.get(reverse("export_session_excel"))
        self.assertEqual(excel_resp.status_code, 403)
        json_resp = self.client.get(reverse("export_session_json"))
        self.assertEqual(json_resp.status_code, 403)

    def test_export_session_excel_contains_grand_data_sheets(self):
        from django.contrib.auth.models import User

        superuser = User.objects.create_superuser(
            username="exportadmin", email="a@a.com", password="password"
        )
        self.client.login(username="exportadmin", password="password")

        response = self.client.get(reverse("export_session_excel"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("session_database_", response["Content-Disposition"])

        from openpyxl import load_workbook

        wb = load_workbook(self._io.BytesIO(response.content))
        for expected_sheet in [
            "Overview", "Sessions", "Terms", "Classes", "Subjects",
            "Departments", "Designations", "Class Subjects", "Parents",
            "Students", "Enrollments", "Subject Assessments", "Mid-Term Records",
            "Grade Verifications", "Staff", "Staff Assignments",
            "Promotion Criteria", "Notifications", "Timetables", "Timetable Slots",
        ]:
            self.assertIn(expected_sheet, wb.sheetnames)

        ws_students = wb["Students"]
        cell_values = [c.value for row in ws_students.iter_rows(min_row=4) for c in row if c.value is not None]
        self.assertIn("EX-010", cell_values)
        self.assertIn("Kwame Mensah", cell_values)

        ws_assessments = wb["Subject Assessments"]
        header_row = [c.value for c in ws_assessments[3]]
        self.assertIn("Subject", header_row)
        cell_values = [c.value for row in ws_assessments.iter_rows(min_row=4) for c in row if c.value is not None]
        self.assertIn("Mathematics", cell_values)
        self.assertIn(90.0, cell_values)

    def test_export_session_json_contains_session_data(self):
        import json
        from django.contrib.auth.models import User

        superuser = User.objects.create_superuser(
            username="exportjsonadmin", email="b@b.com", password="password"
        )
        self.client.login(username="exportjsonadmin", password="password")

        response = self.client.get(reverse("export_session_json"))
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/json")
        self.assertIn("session_database_backup_", response["Content-Disposition"])

        data = json.loads(response.content)
        self.assertEqual(data["session"]["academic_year"], "2025/2026")
        self.assertEqual(data["active_term"]["term_name"], "Term 1")
        self.assertTrue(any(s["admission_number"] == "EX-010" for s in data["students"]))
        assessments = data["subject_assessments"]
        self.assertEqual(len(assessments), 1)
        self.assertEqual(assessments[0]["subject"], "Mathematics")
        self.assertEqual(assessments[0]["total"], 90.0)

    def test_import_requires_superuser(self):
        from django.contrib.auth.models import User

        staff_user = User.objects.create_user(username="importstaff", password="password")
        self.client.login(username="importstaff", password="password")
        response = self.client.post(
            reverse("import_session_json"), {}, format="multipart"
        )
        self.assertEqual(response.status_code, 403)

    def test_import_session_json_restores_records(self):
        import json
        import io
        from django.contrib.auth.models import User

        session = AcademicSession.objects.create(academic_year="2024/2025")
        subject = Subject.objects.create(subject_name="English Language")
        payload = {
            "session": {"academic_year": "2024/2025", "is_current": False},
            "active_term": {"term_name": "Term 2", "is_active": False},
            "sessions": [{"academic_year": "2024/2025", "is_current": False}],
            "terms": [{"session": "2024/2025", "term_name": "Term 2", "is_active": False}],
            "classes": [{"class_name": "JHS 2", "order": 2, "next_class": None, "form_master_id": None}],
            "subjects": ["English Language"],
            "departments": ["Languages"],
            "designations": ["Teacher"],
            "class_subjects": [{"classroom": "JHS 2", "subject": "English Language"}],
            "parents": [{"id": 101, "name": "Ama Owusu", "telephone_number": "0244111222", "occupation": "Trader"}],
            "students": [{
                "admission_number": "IM-001",
                "first_name": "Abena",
                "other_names": "",
                "last_name": "Owusu",
                "gender": "Female",
                "dob": "2011-04-04",
                "date_of_admission": "2024-09-01",
                "status": "Day",
                "living_with": "Mother",
                "classroom": "JHS 2",
                "father_id": None,
                "mother_id": 101,
                "pending_next_class": None,
                "promotion_status": "NEUTRAL",
                "is_alumni": False,
            }],
            "enrollments": [{
                "student": "IM-001",
                "classroom": "JHS 2",
                "academic_session": "2024/2025",
                "academic_term": "Term 2",
            }],
            "subject_assessments": [{
                "student": "IM-001",
                "subject": "English Language",
                "academic_session": "2024/2025",
                "academic_term": "Term 2",
                "class_score": 35.0,
                "exam_score": 45.0,
            }],
        }

        superuser = User.objects.create_superuser(
            username="importadmin", email="c@c.com", password="password"
        )
        self.client.login(username="importadmin", password="password")

        backup = json.dumps(payload).encode('utf-8')
        response = self.client.post(
            reverse("import_session_json"),
            {"backup_file": io.BytesIO(backup)},
            format="multipart",
        )
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("configure_session"))

        self.assertTrue(Student.objects.filter(admission_number="IM-001").exists())
        self.assertTrue(Parent.objects.filter(name="Ama Owusu").exists())
        self.assertTrue(ClassRoom.objects.filter(class_name="JHS 2").exists())
        self.assertTrue(ClassSubject.objects.filter(classroom__class_name="JHS 2").exists())
        self.assertTrue(Subject.objects.filter(subject_name="English Language").exists())
        self.assertTrue(Enrollment.objects.filter(student__admission_number="IM-001").exists())
        self.assertTrue(SubjectAssessment.objects.filter(student__admission_number="IM-001").exists())
        self.assertTrue(Department.objects.filter(name="Languages").exists())
        self.assertTrue(Designation.objects.filter(name="Teacher").exists())

    def test_import_session_json_rejects_foreign_payload(self):
        import json
        import io
        from django.contrib.auth.models import User

        superuser = User.objects.create_superuser(
            username="importbadadmin", email="d@d.com", password="password"
        )
        self.client.login(username="importbadadmin", password="password")

        backup = json.dumps({"foo": "bar"}).encode('utf-8')
        response = self.client.post(
            reverse("import_session_json"),
            {"backup_file": io.BytesIO(backup)},
            format="multipart",
        )
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("configure_session"))
        self.assertEqual(Student.objects.count(), 1)
        self.assertFalse(Student.objects.filter(admission_number="IM-001").exists())

    def test_import_excel_round_trip_restores_exported_workbook(self):
        from openpyxl import Workbook
        import io

        from django.contrib.auth.models import User

        superuser = User.objects.create_superuser(
            username="xlsadmin", email="x@x.com", password="password"
        )
        self.client.login(username="xlsadmin", password="password")

        wb = Workbook()
        ws = wb.active
        ws.title = "Sessions"
        ws.append([])
        ws.append([])
        ws.append(["ID", "Academic Year", "Is Current"])
        ws.append([1, "2025/2026", True])

        wb.create_sheet("Classes")
        ws = wb["Classes"]
        ws.append([])
        ws.append([])
        ws.append(["ID", "Class Name", "Order", "Next Class", "Form Master"])
        ws.append([1, "JHS 1", 1, "", ""])

        wb.create_sheet("Subjects")
        ws = wb["Subjects"]
        ws.append([])
        ws.append([])
        ws.append(["ID", "Subject Name"])
        ws.append([1, "Mathematics"])

        wb.create_sheet("Parents")
        ws = wb["Parents"]
        ws.append([])
        ws.append([])
        ws.append(["ID", "Name", "Occupation", "Residential Address", "Email", "Telephone", "No. of Children"])
        ws.append([101, "Kwame Mensah", "", "", "", "0244000001", 1])

        wb.create_sheet("Students")
        ws = wb["Students"]
        ws.append([])
        ws.append([])
        ws.append([
            "ID", "Admission No.", "First Name", "Other Names", "Last Name", "Gender",
            "Date of Birth", "Date of Admission", "Status", "Living With",
            "Previous School", "Class", "Father Name", "Father Phone", "Father Email",
            "Mother Name", "Mother Phone", "Mother Email",
            "Pending Next Class", "Promotion Status", "Is Alumni",
        ])
        ws.append([
            1, "XLS-001", "Kofi", "", "Mensah", "Male",
            "2011-01-01", "2025-09-01", "Day", "Father",
            "", "JHS 1", "Kwame Mensah", "0244000001", "",
            "", "", "",
            "", "NEUTRAL", False,
        ])

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        response = self.client.post(
            reverse("import_session_excel"),
            {"backup_file": excel_file},
            format="multipart",
        )
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("configure_session"))
        self.assertTrue(Student.objects.filter(admission_number="XLS-001").exists())
        self.assertTrue(ClassRoom.objects.filter(class_name="JHS 1").exists())
        self.assertTrue(Subject.objects.filter(subject_name="Mathematics").exists())
        self.assertTrue(Parent.objects.filter(name="Kwame Mensah").exists())

    def test_import_excel_flexible_maps_custom_headers(self):
        from openpyxl import Workbook
        import io

        from django.contrib.auth.models import User

        superuser = User.objects.create_superuser(
            username="flexadmin", email="y@y.com", password="password"
        )
        self.client.login(username="flexadmin", password="password")

        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        ws.append(["Admission No.", "Student Name", "Class", "Parent Name", "Telephone"])
        ws.append(["FL-001", "Ama Boateng", "JHS 2", "Esi Boateng", "0244111999"])
        ws.append(["FL-002", "Yaw Antwi", "JHS 2", "Yaw Antwi Snr", "0244555666"])

        wb.create_sheet("Classes")
        ws = wb["Classes"]
        ws.append(["Class", "Teacher"])
        ws.append(["JHS 2", "Mrs. Adu"])

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        response = self.client.post(
            reverse("import_session_excel"),
            {"backup_file": excel_file},
            format="multipart",
        )
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("configure_session"))

        self.assertTrue(Student.objects.filter(admission_number="FL-001").exists())
        self.assertTrue(Student.objects.filter(admission_number="FL-002").exists())
        self.assertTrue(ClassRoom.objects.filter(class_name="JHS 2").exists())
        self.assertTrue(Parent.objects.filter(name="Esi Boateng", telephone_number="0244111999").exists())

    def test_import_excel_requires_superuser(self):
        from django.contrib.auth.models import User

        staff_user = User.objects.create_user(username="xlsstaff", password="password")
        self.client.login(username="xlsstaff", password="password")
        response = self.client.post(
            reverse("import_session_excel"), {}, format="multipart"
        )
        self.assertEqual(response.status_code, 403)


class ConfigureSessionTests(TestCase):
    def setUp(self):
        from django.contrib.auth.models import User

        self.user = User.objects.create_superuser(
            username="cfgadmin", password="password", email="cfg@example.com"
        )
        self.client.login(username="cfgadmin", password="password")

    def _make_env(self):
        s1 = AcademicSession.objects.create(academic_year="2024/2025", is_current=True)
        t1a = Term.objects.create(session=s1, term_name="Term 1", is_active=True)
        Term.objects.create(session=s1, term_name="Term 2", is_active=False)
        s2 = AcademicSession.objects.create(academic_year="2025/2026", is_current=False)
        Term.objects.create(session=s2, term_name="Term 1", is_active=False)
        return s1, s2, t1a

    def test_set_active_env_requires_term(self):
        s1, s2, _ = self._make_env()
        response = self.client.post(reverse("configure_session"), {
            "academic_session": s2.pk,
        })
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("configure_session"))
        s1.refresh_from_db()
        s2.refresh_from_db()
        self.assertTrue(s1.is_current)
        self.assertFalse(s2.is_current)
        self.assertEqual(Term.objects.filter(is_active=True).count(), 1)

    def test_set_active_env_rejects_term_from_other_session(self):
        s1, s2, t1a = self._make_env()
        response = self.client.post(reverse("configure_session"), {
            "academic_session": s2.pk,
            "term": t1a.pk,
        })
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("configure_session"))
        s1.refresh_from_db()
        s2.refresh_from_db()
        self.assertTrue(s1.is_current)
        self.assertFalse(s2.is_current)
        self.assertEqual(Term.objects.filter(is_active=True).get(), t1a)

    def test_set_active_env_switches_consistent_pair(self):
        s1, s2, _ = self._make_env()
        s2t1 = Term.objects.get(session=s2, term_name="Term 1")
        response = self.client.post(reverse("configure_session"), {
            "academic_session": s2.pk,
            "term": s2t1.pk,
        })
        self.assertEqual(response.status_code, 302)
        self.assertRedirects(response, reverse("configure_session"))
        s1.refresh_from_db()
        s2.refresh_from_db()
        self.assertFalse(s1.is_current)
        self.assertTrue(s2.is_current)
        self.assertEqual(Term.objects.filter(is_active=True).get(), s2t1)
        self.assertEqual(Term.objects.filter(is_active=True).count(), 1)

    def test_at_most_one_current_session_and_one_active_term(self):
        self._make_env()
        self.assertEqual(AcademicSession.objects.filter(is_current=True).count(), 1)
        self.assertEqual(Term.objects.filter(is_active=True).count(), 1)
